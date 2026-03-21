#!/usr/bin/env python3
"""Generate manual test cases Excel for ClassLite — human-tester perspective.

Every test step is a UI action (click, type, navigate).
Every expected result is visually verifiable on screen.
No database checks, no internal system state, no API inspection.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

TODAY = date.today().isoformat()

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(name="Inter", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Inter", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADERS = [
    "Test Case ID", "Story", "Feature", "Scenario", "Role",
    "Preconditions", "Test Steps", "Expected Result",
    "Priority", "Regression", "Last Updated",
    "Status", "Tester", "Notes",
]

COL_WIDTHS = [14, 8, 22, 40, 10, 30, 50, 45, 8, 10, 14, 10, 10, 20]


def add_sheet(wb, name, cases):
    ws = wb.create_sheet(title=name)
    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    # Data rows
    for row_idx, case in enumerate(cases, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            val = case.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}{len(cases)+1}"
    return ws


def c(id, story, feature, scenario, role, preconditions, steps, expected, priority="P1", regression=""):
    """Helper to create a test case dict."""
    return {
        "Test Case ID": id,
        "Story": story,
        "Feature": feature,
        "Scenario": scenario,
        "Role": role,
        "Preconditions": preconditions,
        "Test Steps": steps,
        "Expected Result": expected,
        "Priority": priority,
        "Regression": regression,
        "Last Updated": TODAY,
    }


# ──────────────────────────────────────────────
# EPIC 1 — Tenant & Users
# ──────────────────────────────────────────────
def epic1():
    cases = []
    # Story 1.1 — Multi-Tenant Onboarding
    cases.append(c("1.1-001", "1.1", "Center Registration", "Register new center via Google OAuth",
        "OWNER", "No existing account; Google account available",
        "1. Navigate to /sign-up/center\n2. Enter a center name (e.g., 'IELTS Academy')\n3. Click 'Sign up with Google'\n4. Select your Google account in the popup\n5. Wait for redirect",
        "1. You are redirected to the center dashboard\n2. The center name you entered appears in the sidebar/nav rail\n3. Your Google profile name appears in the user menu (top-right)",
        "P1", "Yes"))
    cases.append(c("1.1-002", "1.1", "Center Registration", "Register new center via email/password",
        "OWNER", "No existing account",
        "1. Navigate to /sign-up/center\n2. Enter center name, your name, email, and password\n3. Click the 'Register' button\n4. Wait for redirect",
        "1. You are redirected to the center dashboard\n2. The center name appears in the sidebar\n3. Your name appears in the user menu",
        "P1", "Yes"))
    cases.append(c("1.1-003", "1.1", "Center Registration", "Registration fails with duplicate email",
        "OWNER", "An account already exists with the same email",
        "1. Navigate to /sign-up/center\n2. Enter a center name and use an email that already has an account\n3. Fill in remaining fields and submit",
        "An error message appears indicating the email is already in use. No redirect occurs.",
        "P1", "Yes"))
    cases.append(c("1.1-004", "1.1", "Data Isolation", "Center A data is never visible to Center B",
        "OWNER", "Two centers exist, each with courses, classes, and students",
        "1. Log in as Owner of Center A\n2. Navigate to Classes, Exercises, Students, Assignments, and Schedule pages\n3. Note all data visible\n4. Log out\n5. Log in as Owner of Center B\n6. Navigate to the same pages",
        "Center B sees only its own data on every page. No courses, classes, students, or exercises from Center A appear.",
        "P1", "Yes"))
    cases.append(c("1.1-005", "1.1", "Data Isolation", "URL manipulation cannot access other center's data",
        "OWNER", "Logged in as Owner of Center A; know Center B's centerId",
        "1. While logged into Center A, manually change the centerId in the browser URL to Center B's ID\n2. Try to load a page (e.g., /centerB-id/dashboard/classes)",
        "You are either redirected back to your own center's dashboard or see an unauthorized/forbidden error. No Center B data is displayed.",
        "P1", "Yes"))

    # Story 1.2 — Center Branding
    cases.append(c("1.2-001", "1.2", "Center Branding", "Upload a valid PNG logo under 2MB",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > General\n2. Click the logo upload area or 'Upload Logo' button\n3. Select a PNG file smaller than 2MB\n4. Wait for upload to complete",
        "1. The uploaded logo appears in the Settings page preview\n2. The logo appears in the sidebar/nav rail\n3. All other users see the new logo after refreshing",
        "P1", ""))
    cases.append(c("1.2-002", "1.2", "Center Branding", "Reject logo file over 2MB",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > General\n2. Click the logo upload area\n3. Select an image file larger than 2MB",
        "An error message appears indicating the file exceeds the 2MB limit. The logo is not updated.",
        "P1", ""))
    cases.append(c("1.2-003", "1.2", "Center Branding", "Update center name",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > General\n2. Edit the center name field\n3. Click Save",
        "1. Success toast appears\n2. The new center name appears in the sidebar/nav rail\n3. Refreshing the page shows the new name persisted",
        "P1", ""))
    cases.append(c("1.2-004", "1.2", "Center Branding", "Set center timezone",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > General\n2. Change the timezone dropdown to a different timezone\n3. Click Save",
        "1. Success toast appears\n2. Schedule page shows times in the new timezone",
        "P2", ""))
    cases.append(c("1.2-005", "1.2", "Center Branding", "Non-owner cannot access General Settings",
        "TEACHER", "Logged in as Teacher",
        "1. Navigate to the sidebar\n2. Confirm there is no 'Settings' link\n3. Manually enter the Settings > General URL in the browser",
        "Settings link is not visible in the sidebar. Direct URL access redirects to the dashboard or shows an unauthorized message.",
        "P1", "Yes"))

    # Story 1.3 — User Invitation & RBAC
    cases.append(c("1.3-001", "1.3", "User Invitation", "Invite a Teacher via email",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > Users\n2. Click 'Invite User'\n3. Enter a valid email address\n4. Select role 'Teacher' from dropdown\n5. Click Send / Invite",
        "1. Success toast appears (e.g., 'Invitation sent')\n2. The invited email appears in the 'Pending Invitations' tab with status 'Pending'",
        "P1", "Yes"))
    cases.append(c("1.3-002", "1.3", "User Invitation", "Invite a Student via email",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > Users\n2. Click 'Invite User'\n3. Enter a valid email\n4. Select role 'Student'\n5. Click Send",
        "1. Success toast appears\n2. Invitation appears in 'Pending Invitations' tab",
        "P1", "Yes"))
    cases.append(c("1.3-003", "1.3", "User Invitation", "Invite an Admin via email",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > Users\n2. Click 'Invite User'\n3. Enter a valid email\n4. Select role 'Admin'\n5. Click Send",
        "1. Success toast appears\n2. Invitation appears in 'Pending Invitations' tab",
        "P1", ""))
    cases.append(c("1.3-004", "1.3", "User Invitation", "Invited user signs up via invitation link",
        "TEACHER", "Invitation email received with sign-up link",
        "1. Open the invitation email\n2. Click the sign-up link\n3. Complete the registration form (name, password)\n4. Submit",
        "1. User is redirected to the center's dashboard\n2. User appears in Settings > Users as 'Active' with the correct role\n3. The invitation in 'Pending Invitations' disappears",
        "P1", "Yes"))
    cases.append(c("1.3-005", "1.3", "User Invitation", "Resend a pending invitation",
        "OWNER", "A pending invitation exists",
        "1. Navigate to Settings > Users > Pending Invitations tab\n2. Find the pending invitation\n3. Click 'Resend'",
        "1. Success toast appears\n2. The invitation's expiry date is updated",
        "P2", ""))
    cases.append(c("1.3-006", "1.3", "User Invitation", "Revoke a pending invitation",
        "OWNER", "A pending invitation exists",
        "1. Navigate to Settings > Users > Pending Invitations tab\n2. Find the pending invitation\n3. Click 'Revoke'",
        "1. Confirmation dialog appears\n2. After confirming, the invitation is removed from the list\n3. The invitation link no longer works (invitee sees error if they try to use it)",
        "P1", ""))
    cases.append(c("1.3-007", "1.3", "User Invitation", "Duplicate invitation to same email shows error",
        "OWNER", "A pending invitation already exists for test@example.com",
        "1. Navigate to Settings > Users\n2. Click 'Invite User'\n3. Enter test@example.com (same as existing invitation)\n4. Select any role and send",
        "An error message appears indicating an invitation already exists for this email.",
        "P2", ""))

    # Story 1.4 — RBAC UI
    cases.append(c("1.4-001", "1.4", "RBAC UI", "Student cannot see management nav items",
        "STUDENT", "Logged in as Student",
        "1. Look at the sidebar navigation\n2. Note which items are visible",
        "Student sees only: Dashboard, Schedule, and My Profile. Does NOT see: Classes, Exercises, Assignments, Mock Tests, Grading, Students, or Settings.",
        "P1", "Yes"))
    cases.append(c("1.4-002", "1.4", "RBAC UI", "Teacher sees management items but not Settings",
        "TEACHER", "Logged in as Teacher",
        "1. Look at the sidebar navigation\n2. Note which items are visible",
        "Teacher sees: Dashboard, Schedule, Classes, Exercises, Assignments, Mock Tests, Grading, Students, My Profile. Does NOT see: Settings.",
        "P1", "Yes"))
    cases.append(c("1.4-003", "1.4", "RBAC UI", "Owner sees all navigation items including Settings",
        "OWNER", "Logged in as Owner",
        "1. Look at the sidebar navigation\n2. Note which items are visible",
        "Owner sees all items: Dashboard, Schedule, Classes, Exercises, Assignments, Mock Tests, Grading, Students, Settings, My Profile.",
        "P1", "Yes"))
    cases.append(c("1.4-004", "1.4", "RBAC UI", "Admin sees all navigation items including Settings",
        "ADMIN", "Logged in as Admin",
        "1. Look at the sidebar navigation\n2. Note which items are visible",
        "Admin sees the same navigation items as Owner.",
        "P1", "Yes"))
    cases.append(c("1.4-005", "1.4", "RBAC UI", "Student cannot access Exercises page via direct URL",
        "STUDENT", "Logged in as Student",
        "1. Manually type /:centerId/dashboard/exercises in the browser address bar\n2. Press Enter",
        "The page redirects to the student dashboard or shows an 'Access Denied' message. The Exercises page does not load.",
        "P1", "Yes"))

    # Story 1.5 — Dashboard Shell
    cases.append(c("1.5-001", "1.5", "Dashboard Shell", "Owner dashboard shows Student Health view",
        "OWNER", "Logged in as Owner; center has students with varied attendance",
        "1. Navigate to the Dashboard (click Dashboard in sidebar)\n2. Observe the main content area",
        "Dashboard shows the Student Health view with student cards color-coded by status (red/yellow/green). Health summary bar with status filters is visible.",
        "P1", "Yes"))
    cases.append(c("1.5-002", "1.5", "Dashboard Shell", "Teacher dashboard shows At Risk widget and Grading queue",
        "TEACHER", "Logged in as Teacher; has assigned classes",
        "1. Navigate to the Dashboard",
        "Dashboard shows an 'At Risk' students widget and a Grading Queue summary card.",
        "P1", "Yes"))
    cases.append(c("1.5-003", "1.5", "Dashboard Shell", "Student dashboard shows assignments grouped by urgency",
        "STUDENT", "Logged in as Student; has assignments with various due dates",
        "1. Navigate to the Dashboard",
        "Dashboard shows assignment cards grouped by: Overdue, Due Today, Due This Week, Upcoming, No Deadline. Cards show assignment title, due date, and status badge.",
        "P1", "Yes"))
    cases.append(c("1.5-004", "1.5", "Dashboard Shell", "Sidebar collapses to bottom bar on mobile",
        "ALL", "Logged in; using mobile device or browser window resized to <768px width",
        "1. Resize browser window to <768px wide (or use mobile device)\n2. Observe the navigation",
        "1. Sidebar disappears\n2. A bottom navigation bar appears with 4 main icons: Dashboard, Schedule, Classes, Exercises\n3. A 'More' button reveals remaining nav items in an overlay",
        "P1", "Yes"))
    cases.append(c("1.5-005", "1.5", "Dashboard Shell", "Center logo and name display in sidebar",
        "OWNER", "Center has a logo and name configured",
        "1. Log in and navigate to the Dashboard\n2. Look at the top of the sidebar",
        "The center's logo and name appear at the top of the sidebar navigation.",
        "P1", ""))

    # Story 1.6 — Login & Session
    cases.append(c("1.6-001", "1.6", "Login", "Login with valid email and password",
        "ALL", "User account exists with email/password",
        "1. Navigate to /sign-in\n2. Enter valid email and password\n3. Click 'Sign In'",
        "1. Brief loading indicator appears on button\n2. You are redirected to the dashboard\n3. Your name appears in the user menu (top-right)",
        "P1", "Yes"))
    cases.append(c("1.6-002", "1.6", "Login", "Login with Google OAuth",
        "ALL", "User account exists and is linked to Google",
        "1. Navigate to /sign-in\n2. Click 'Sign in with Google'\n3. Select your Google account in the popup",
        "1. You are redirected to the dashboard\n2. Your Google profile name appears in the user menu",
        "P1", "Yes"))
    cases.append(c("1.6-003", "1.6", "Login", "Login with invalid password shows error",
        "ALL", "User account exists",
        "1. Navigate to /sign-in\n2. Enter valid email but wrong password\n3. Click 'Sign In'",
        "An error message appears (e.g., 'Invalid email or password'). You remain on the login page.",
        "P1", "Yes"))
    cases.append(c("1.6-004", "1.6", "Login", "Account lockout after 5 failed attempts",
        "ALL", "User account exists",
        "1. Navigate to /sign-in\n2. Enter valid email but wrong password\n3. Click 'Sign In'\n4. Repeat steps 2-3 four more times (5 total failed attempts)",
        "After the 5th failed attempt, an error message appears indicating the account is temporarily locked (e.g., 'Too many failed attempts. Try again in 15 minutes').",
        "P1", "Yes"))
    cases.append(c("1.6-005", "1.6", "Login", "Log out clears session",
        "ALL", "Logged in",
        "1. Click the user menu (top-right avatar/name)\n2. Click 'Log out'\n3. Try to navigate back to the dashboard URL",
        "1. You are redirected to the login page\n2. Navigating to the dashboard URL redirects back to login\n3. No user data is visible",
        "P1", "Yes"))
    cases.append(c("1.6-006", "1.6", "Login", "Session expires and redirects to login",
        "ALL", "Logged in; session has expired (wait or simulate expiry)",
        "1. Leave the app idle until session expires\n2. Try to navigate to a page or perform an action",
        "You are redirected to /sign-in?expired=true with a message like 'Your session has expired. Please sign in again.'",
        "P2", ""))
    cases.append(c("1.6-007", "1.6", "Login", "Remember Me extends session",
        "ALL", "On the login page",
        "1. Navigate to /sign-in\n2. Enter valid email and password\n3. Check the 'Remember me' checkbox\n4. Click 'Sign In'\n5. Close the browser completely\n6. Reopen and navigate to the app",
        "You are still logged in without needing to sign in again.",
        "P2", ""))

    # Story 1.7 — Password Reset
    cases.append(c("1.7-001", "1.7", "Password Reset", "Request password reset email",
        "ALL", "User account exists with email/password",
        "1. Navigate to /sign-in\n2. Click 'Forgot password?'\n3. Enter your email address\n4. Click 'Send Reset Link'",
        "1. A success message appears: 'If an account exists with this email, a reset link has been sent'\n2. Check email inbox for the reset link",
        "P1", "Yes"))
    cases.append(c("1.7-002", "1.7", "Password Reset", "Reset password with valid link",
        "ALL", "Password reset email received",
        "1. Click the reset link in the email\n2. Enter a new password (8+ chars, 1 uppercase, 1 number)\n3. Confirm the new password\n4. Click 'Reset Password'",
        "1. Success message appears\n2. You are redirected to /sign-in with a toast 'Password has been reset'\n3. You can log in with the new password",
        "P1", "Yes"))
    cases.append(c("1.7-003", "1.7", "Password Reset", "Reset link expires after 1 hour",
        "ALL", "Password reset was requested more than 1 hour ago",
        "1. Click the expired reset link from the email",
        "An error message appears indicating the link has expired, with an option to request a new one.",
        "P2", ""))
    cases.append(c("1.7-004", "1.7", "Password Reset", "Weak password rejected during reset",
        "ALL", "On the password reset page with valid link",
        "1. Enter a weak password (e.g., '1234')\n2. Click 'Reset Password'",
        "Validation error appears indicating password requirements (8+ characters, 1 uppercase, 1 number).",
        "P1", ""))
    cases.append(c("1.7-005", "1.7", "Password Reset", "Non-existent email shows same success message (no enumeration)",
        "ALL", "",
        "1. Navigate to /forgot-password\n2. Enter an email that has no account\n3. Click 'Send Reset Link'",
        "The same generic success message appears ('If an account exists...'). No indication of whether the email exists.",
        "P1", ""))

    # Story 1.8 — User Management
    cases.append(c("1.8-001", "1.8", "User Management", "View paginated user list",
        "OWNER", "Center has 25+ users",
        "1. Navigate to Settings > Users\n2. Observe the user table",
        "User table displays with columns: Name, Email, Role (color badge), Status, Last Active, Actions. Pagination controls appear at the bottom.",
        "P1", "Yes"))
    cases.append(c("1.8-002", "1.8", "User Management", "Search users by name",
        "OWNER", "Center has multiple users",
        "1. Navigate to Settings > Users\n2. Type a user's name in the search field",
        "The table filters to show only users whose name matches the search text.",
        "P1", ""))
    cases.append(c("1.8-003", "1.8", "User Management", "Filter users by role",
        "OWNER", "Center has users with different roles",
        "1. Navigate to Settings > Users\n2. Select 'Teacher' from the role filter dropdown",
        "The table shows only users with the Teacher role.",
        "P1", ""))
    cases.append(c("1.8-004", "1.8", "User Management", "Change a user's role (Owner only)",
        "OWNER", "Center has at least one Teacher user",
        "1. Navigate to Settings > Users\n2. Find a Teacher in the list\n3. Click the Actions menu for that user\n4. Select 'Change Role'\n5. Select 'Admin' from the dropdown\n6. Confirm the change",
        "1. Confirmation dialog explains the permission changes\n2. After confirming, the user's role badge updates to 'Admin'\n3. Success toast appears",
        "P1", "Yes"))
    cases.append(c("1.8-005", "1.8", "User Management", "Admin cannot change user roles",
        "ADMIN", "Logged in as Admin",
        "1. Navigate to Settings > Users\n2. Find a Teacher in the list\n3. Check the Actions menu",
        "The 'Change Role' option is not available or is disabled in the actions menu.",
        "P1", "Yes"))
    cases.append(c("1.8-006", "1.8", "User Management", "Deactivate a user",
        "OWNER", "Center has an active non-Owner user",
        "1. Navigate to Settings > Users\n2. Find the user\n3. Click Actions > Deactivate\n4. Confirm",
        "1. User's status changes to 'Inactive'\n2. The deactivated user can no longer log in (verify by trying)",
        "P1", "Yes"))
    cases.append(c("1.8-007", "1.8", "User Management", "Reactivate a deactivated user",
        "OWNER", "A deactivated user exists",
        "1. Navigate to Settings > Users\n2. Filter by 'Inactive' status\n3. Find the deactivated user\n4. Click Actions > Reactivate",
        "1. User's status changes to 'Active'\n2. The user can log in again",
        "P2", ""))
    cases.append(c("1.8-008", "1.8", "User Management", "View Pending Invitations tab",
        "OWNER", "Invitations have been sent",
        "1. Navigate to Settings > Users\n2. Click the 'Pending Invitations' tab",
        "A list of pending invitations appears showing email, role, invited date, and expiry status. 'Resend' and 'Revoke' action buttons are visible per invitation.",
        "P1", ""))
    cases.append(c("1.8-009", "1.8", "User Management", "Delete a user",
        "OWNER", "Center has a non-Owner user",
        "1. Navigate to Settings > Users\n2. Find the user\n3. Click Actions > Delete\n4. Confirm in the dialog",
        "1. Confirmation dialog warns about permanent deletion\n2. After confirming, the user is removed from the list\n3. Success toast appears",
        "P1", ""))
    cases.append(c("1.8-010", "1.8", "User Management", "Bulk deactivate multiple users",
        "OWNER", "Center has multiple active non-Owner users",
        "1. Navigate to Settings > Users\n2. Select checkboxes for 2-3 users\n3. Click the bulk 'Deactivate' action\n4. Confirm",
        "All selected users' statuses change to 'Inactive'. Success toast appears.",
        "P2", ""))

    # Story 1.9 — Profile Self-Service
    cases.append(c("1.9-001", "1.9", "Profile", "Update display name",
        "ALL", "Logged in",
        "1. Navigate to My Profile\n2. Click Edit (or the edit icon)\n3. Change the Display Name\n4. Click Save",
        "1. Success toast appears\n2. The new name is reflected on the profile page\n3. The user menu (top-right) shows the updated name",
        "P1", "Yes"))
    cases.append(c("1.9-002", "1.9", "Profile", "Upload profile photo",
        "ALL", "Logged in",
        "1. Navigate to My Profile\n2. Click the avatar/photo area\n3. Select an image file under 1MB\n4. Wait for upload",
        "1. The new photo appears on the profile page\n2. The avatar in the user menu and sidebar updates",
        "P1", ""))
    cases.append(c("1.9-003", "1.9", "Profile", "Reject profile photo over 1MB",
        "ALL", "Logged in",
        "1. Navigate to My Profile\n2. Click the avatar area\n3. Select an image file larger than 1MB",
        "Error message indicates the file exceeds the size limit. The photo is not updated.",
        "P2", ""))
    cases.append(c("1.9-004", "1.9", "Profile", "Email field is read-only",
        "ALL", "Logged in",
        "1. Navigate to My Profile\n2. Try to click on or edit the Email field",
        "The email field is not editable. A tooltip or note explains why (e.g., 'Contact admin to change email').",
        "P1", ""))
    cases.append(c("1.9-005", "1.9", "Profile", "Role field is read-only",
        "ALL", "Logged in",
        "1. Navigate to My Profile\n2. Look at the Role field",
        "The role is displayed but not editable. A tooltip explains that role changes require an admin.",
        "P1", ""))
    cases.append(c("1.9-006", "1.9", "Profile", "Change password (email/password user)",
        "ALL", "Logged in with email/password provider",
        "1. Navigate to My Profile\n2. Scroll to 'Change Password' section\n3. Enter current password\n4. Enter new password (meeting requirements)\n5. Confirm new password\n6. Click 'Change Password'",
        "1. Success toast appears\n2. You are logged out of other sessions\n3. You can log in with the new password",
        "P1", "Yes"))
    cases.append(c("1.9-007", "1.9", "Profile", "Change password unavailable for Google OAuth users",
        "ALL", "Logged in via Google OAuth",
        "1. Navigate to My Profile\n2. Look for 'Change Password' section",
        "The change password section is hidden or shows a message: 'Password change is not available for Google sign-in accounts.'",
        "P1", ""))
    cases.append(c("1.9-008", "1.9", "Profile", "Request account deletion (non-Owner)",
        "TEACHER", "Logged in as Teacher",
        "1. Navigate to My Profile\n2. Scroll to bottom\n3. Click 'Delete My Account'\n4. Confirm in the dialog",
        "1. An alert banner appears on the profile showing 'Account scheduled for deletion' with a countdown (7-day grace period)\n2. A 'Cancel Deletion' option is available",
        "P1", ""))
    cases.append(c("1.9-009", "1.9", "Profile", "Cancel pending account deletion",
        "TEACHER", "Account deletion has been requested (within 7-day window)",
        "1. Navigate to My Profile\n2. See the deletion pending banner\n3. Click 'Cancel Deletion'",
        "1. The banner disappears\n2. Account returns to normal active state",
        "P1", ""))
    cases.append(c("1.9-010", "1.9", "Profile", "Owner cannot delete own account",
        "OWNER", "Logged in as Owner",
        "1. Navigate to My Profile\n2. Look for 'Delete My Account' button",
        "The 'Delete My Account' button is not visible for the Owner role.",
        "P1", ""))
    cases.append(c("1.9-011", "1.9", "Profile", "Update phone number",
        "ALL", "Logged in",
        "1. Navigate to My Profile\n2. Click Edit\n3. Enter or update the phone number\n4. Click Save",
        "1. Success toast appears\n2. Phone number is saved and visible on profile",
        "P2", ""))

    # Story 1.10 — CSV Bulk Import
    cases.append(c("1.10-001", "1.10", "CSV Import", "Download CSV template",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > Users\n2. Click 'Import CSV' or 'CSV Import' button\n3. Click 'Download Template'",
        "A CSV file downloads with columns: Email, Name, Role (and optionally sample data rows).",
        "P1", ""))
    cases.append(c("1.10-002", "1.10", "CSV Import", "Upload valid CSV and preview rows",
        "OWNER", "CSV template filled with 5 valid rows",
        "1. Click 'Import CSV'\n2. Upload the filled CSV file\n3. Observe the preview",
        "1. A preview table shows the first 5 rows with Email, Name, Role columns\n2. A summary shows: 'X valid rows, 0 errors'\n3. A 'Confirm Import' button is available",
        "P1", "Yes"))
    cases.append(c("1.10-003", "1.10", "CSV Import", "CSV with duplicate emails shows validation errors",
        "OWNER", "CSV contains emails that already exist in the center",
        "1. Click 'Import CSV'\n2. Upload the CSV with duplicate emails",
        "1. The preview shows error indicators on duplicate rows\n2. Summary shows: 'X valid, Y duplicates'\n3. Duplicate rows can be deselected before importing",
        "P1", ""))
    cases.append(c("1.10-004", "1.10", "CSV Import", "Confirm import creates users",
        "OWNER", "CSV uploaded with valid rows previewed",
        "1. After previewing, click 'Confirm Import'\n2. Wait for processing to complete",
        "1. Progress indicator shows import progress\n2. After completion, success message shows how many users were created\n3. New users appear in the Users list with 'Active' status",
        "P1", "Yes"))
    cases.append(c("1.10-005", "1.10", "CSV Import", "Reject CSV over 5MB",
        "OWNER", "A CSV file larger than 5MB",
        "1. Click 'Import CSV'\n2. Try to upload the large CSV file",
        "An error message appears indicating the file exceeds the 5MB limit.",
        "P2", ""))
    cases.append(c("1.10-006", "1.10", "CSV Import", "Reject CSV with more than 1000 rows",
        "OWNER", "A CSV file with over 1000 data rows",
        "1. Click 'Import CSV'\n2. Upload the CSV",
        "An error message indicates the maximum of 1,000 rows has been exceeded.",
        "P2", ""))
    cases.append(c("1.10-007", "1.10", "CSV Import", "View import history",
        "OWNER", "At least one CSV import has been completed",
        "1. Navigate to Settings > Users\n2. Click the 'Import History' tab",
        "A table shows past imports with: Date, File name, Total rows, Success count, Error count, Status.",
        "P1", ""))

    # Story 1.11 — Navigation
    cases.append(c("1.11-001", "1.11", "Navigation", "Sidebar highlights active page",
        "ALL", "Logged in",
        "1. Click 'Dashboard' in the sidebar\n2. Observe the sidebar\n3. Click 'Schedule'\n4. Observe the sidebar again",
        "The active navigation item is highlighted (blue background with white text). When you switch pages, the highlight moves to the new page.",
        "P1", ""))
    cases.append(c("1.11-002", "1.11", "Navigation", "Breadcrumbs show current location",
        "OWNER", "Logged in",
        "1. Navigate to Settings > Users\n2. Look at the top bar area",
        "Breadcrumbs show: 'Settings > Users'. Each breadcrumb segment is clickable to navigate back.",
        "P2", ""))
    cases.append(c("1.11-003", "1.11", "Navigation", "Mobile bottom bar shows overflow menu",
        "ALL", "Using mobile device or window <768px",
        "1. Look at the bottom navigation bar\n2. Tap the 'More' icon (rightmost)",
        "An overlay menu appears showing remaining navigation items not visible in the bottom bar.",
        "P1", ""))
    cases.append(c("1.11-004", "1.11", "Navigation", "Theme toggle switches light/dark mode",
        "ALL", "Logged in",
        "1. Click the theme toggle icon in the top bar\n2. Observe the page",
        "The entire UI switches between light and dark mode. Colors, backgrounds, and text adapt accordingly.",
        "P2", ""))
    cases.append(c("1.11-005", "1.11", "Navigation", "Teacher sees '(Read)' badge on Classes nav item",
        "TEACHER", "Logged in as Teacher",
        "1. Look at the 'Classes' item in the sidebar",
        "The Classes nav item shows a '(Read)' badge indicating read-only access.",
        "P2", ""))
    cases.append(c("1.11-006", "1.11", "Navigation", "Settings tabs: General, Users, Rooms, Tags, Integrations, Privacy, Billing",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings\n2. Observe the tab/vertical navigation within Settings",
        "The following tabs are visible: General, Users, Rooms, Tags, Integrations, Privacy, Billing. Clicking each tab loads the corresponding settings page.",
        "P1", ""))

    # Permission boundary tests
    cases.append(c("1.4-006", "1.4", "Permission Boundaries", "Student cannot access Grading page",
        "STUDENT", "Logged in as Student",
        "1. Manually enter /:centerId/dashboard/grading in the browser URL",
        "Redirected to dashboard or shown 'Access Denied'. Grading page does not load.",
        "P1", "Yes"))
    cases.append(c("1.4-007", "1.4", "Permission Boundaries", "Student cannot access Settings page",
        "STUDENT", "Logged in as Student",
        "1. Manually enter /:centerId/dashboard/settings in the browser URL",
        "Redirected to dashboard or shown 'Access Denied'. Settings page does not load.",
        "P1", "Yes"))
    cases.append(c("1.4-008", "1.4", "Permission Boundaries", "Teacher cannot access Billing settings",
        "TEACHER", "Logged in as Teacher",
        "1. Manually enter /:centerId/dashboard/settings/billing in the browser URL",
        "Redirected to dashboard or shown 'Access Denied'. Billing page does not load.",
        "P1", "Yes"))
    cases.append(c("1.4-009", "1.4", "Permission Boundaries", "Student cannot access Students health page",
        "STUDENT", "Logged in as Student",
        "1. Manually enter /:centerId/dashboard/students in the browser URL",
        "Redirected to dashboard or shown 'Access Denied'.",
        "P1", "Yes"))
    cases.append(c("1.4-010", "1.4", "Permission Boundaries", "Teacher cannot access Settings > Users",
        "TEACHER", "Logged in as Teacher",
        "1. Manually enter /:centerId/dashboard/settings/users in the browser URL",
        "Redirected to dashboard or shown 'Access Denied'. User management page does not load.",
        "P1", "Yes"))

    return cases


# ──────────────────────────────────────────────
# EPIC 2 — Scheduling
# ──────────────────────────────────────────────
def epic2():
    cases = []
    # Story 2.1 — Course & Class Management
    cases.append(c("2.1-001", "2.1", "Course Management", "Create a new course",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Classes page\n2. Click the 'Courses' tab\n3. Click '+ New Course'\n4. In the drawer, enter: Name (e.g., 'IELTS Foundation'), Description, and select a Brand Color\n5. Click Save",
        "1. Success toast appears\n2. The course drawer closes\n3. The new course appears in the Courses table with the entered name, description, and color swatch",
        "P1", "Yes"))
    cases.append(c("2.1-002", "2.1", "Course Management", "Edit an existing course",
        "OWNER", "At least one course exists",
        "1. Navigate to Classes > Courses tab\n2. Click the Edit (pencil) icon on a course\n3. Modify the name and/or description\n4. Click Save",
        "1. Success toast appears\n2. The updated values appear in the Courses table",
        "P1", ""))
    cases.append(c("2.1-003", "2.1", "Course Management", "Delete a course",
        "OWNER", "A course exists with no linked classes (or test with linked classes)",
        "1. Navigate to Classes > Courses tab\n2. Click the Delete (trash) icon on a course\n3. Confirm in the dialog",
        "1. Confirmation dialog warns about deletion\n2. After confirming, the course is removed from the table\n3. Success toast appears",
        "P1", ""))
    cases.append(c("2.1-004", "2.1", "Course Management", "Change course brand color",
        "OWNER", "A course exists",
        "1. Edit a course\n2. Change the Brand Color to a different color\n3. Save\n4. Navigate to the Schedule page",
        "1. The color swatch in the Courses table updates\n2. On the Schedule page, sessions for this course use the new color",
        "P2", ""))
    cases.append(c("2.1-005", "2.1", "Class Management", "Create a new class linked to a course",
        "OWNER", "At least one course exists",
        "1. Navigate to Classes page > Classes tab\n2. Click '+ New Class'\n3. Enter class name (e.g., 'Class 10A')\n4. Select a course from the dropdown\n5. Click Save",
        "1. Success toast appears\n2. The new class appears in the Classes table with the class name and linked course name",
        "P1", "Yes"))
    cases.append(c("2.1-006", "2.1", "Class Management", "Assign a teacher to a class",
        "OWNER", "A class and at least one Teacher user exist",
        "1. Click the Edit icon on a class\n2. In the drawer, select a Teacher from the teacher dropdown\n3. Click Save",
        "1. Success toast appears\n2. The teacher name appears in the class details/table",
        "P1", ""))
    cases.append(c("2.1-007", "2.1", "Class Management", "Add students to class roster",
        "OWNER", "A class exists; Student users exist",
        "1. Click Edit on a class\n2. In the drawer, navigate to the roster section\n3. Search for a student by name\n4. Add the student to the roster\n5. Save",
        "1. The student appears in the class roster list\n2. Student count in the Classes table updates",
        "P1", "Yes"))
    cases.append(c("2.1-008", "2.1", "Class Management", "Remove a student from class roster",
        "OWNER", "A class with students enrolled",
        "1. Edit the class\n2. In the roster section, click remove/delete on a student\n3. Save",
        "1. The student is removed from the roster\n2. Student count decreases",
        "P1", ""))
    cases.append(c("2.1-009", "2.1", "Class Management", "Teacher has read-only access to classes",
        "TEACHER", "Logged in as Teacher assigned to a class",
        "1. Navigate to Classes page\n2. Try to edit or delete a class",
        "The Edit and Delete actions are not available or are disabled. Teacher can view class details but cannot modify them.",
        "P1", "Yes"))
    cases.append(c("2.1-010", "2.1", "Class Management", "Delete a class",
        "OWNER", "A class exists",
        "1. Navigate to Classes > Classes tab\n2. Click Delete on a class\n3. Confirm",
        "1. Confirmation dialog appears\n2. Class is removed from the table after confirming\n3. Success toast appears",
        "P1", ""))
    cases.append(c("2.1-011", "2.1", "Class Management", "Filter class roster by student name",
        "OWNER", "A class with multiple students",
        "1. Edit a class to open the drawer\n2. In the roster section, type a student's name in the search field",
        "The roster list filters to show only students matching the search text.",
        "P2", ""))
    cases.append(c("2.1-012", "2.1", "Class Management", "Set default room for a class",
        "OWNER", "A class exists; rooms configured in Settings",
        "1. Edit a class\n2. Select a default room from the Room dropdown\n3. Save",
        "The default room is saved and pre-selected when creating sessions for this class.",
        "P2", ""))

    # Story 2.2 — Visual Weekly Scheduler
    cases.append(c("2.2-001", "2.2", "Weekly Scheduler", "View weekly calendar with session blocks",
        "OWNER", "Sessions exist for the current week",
        "1. Navigate to the Schedule page",
        "1. A 7-column grid (Mon-Sun) displays with time slots from 8 AM to 10 PM\n2. Existing sessions appear as colored blocks using their course's brand color\n3. Each block shows the course/class name",
        "P1", "Yes"))
    cases.append(c("2.2-002", "2.2", "Weekly Scheduler", "Navigate between weeks",
        "ALL", "Logged in",
        "1. Navigate to the Schedule page\n2. Click the 'Next' arrow button\n3. Click the 'Previous' arrow button\n4. Click 'Today'",
        "1. Next: calendar shifts to the following week\n2. Previous: calendar shifts to the prior week\n3. Today: calendar returns to the current week",
        "P1", ""))
    cases.append(c("2.2-003", "2.2", "Weekly Scheduler", "Click session block to view details",
        "ALL", "Sessions exist",
        "1. Click on a session block in the calendar",
        "A popover or detail panel shows: Course Name, Class Name, Room/Location, Teacher Name, Roster Size (number of students).",
        "P1", ""))
    cases.append(c("2.2-004", "2.2", "Weekly Scheduler", "Drag-and-drop session to reschedule (Owner/Admin)",
        "OWNER", "A session exists",
        "1. Click and hold a session block\n2. Drag it to a different time slot or day\n3. Release",
        "1. The session block moves to the new position\n2. The session's date/time is updated\n3. If a conflict exists, a warning banner appears (see conflict detection tests)",
        "P1", "Yes"))
    cases.append(c("2.2-005", "2.2", "Weekly Scheduler", "Teacher/Student cannot drag sessions",
        "TEACHER", "Sessions exist",
        "1. Try to click and drag a session block to a different time",
        "The session block does not move. Drag-and-drop is disabled for Teacher and Student roles.",
        "P1", "Yes"))
    cases.append(c("2.2-006", "2.2", "Weekly Scheduler", "Mobile: calendar collapses to single-day view",
        "ALL", "Using mobile device or window <768px",
        "1. Open the Schedule page on a mobile device\n2. Observe the layout",
        "1. Calendar shows a single day at a time\n2. A horizontal date picker appears at the top for switching days\n3. Session blocks are shown for the selected day only",
        "P1", "Yes"))
    cases.append(c("2.2-007", "2.2", "Weekly Scheduler", "Session uses course brand color",
        "ALL", "A course with a specific brand color has sessions",
        "1. Navigate to Schedule\n2. Observe the session blocks",
        "Session blocks for that course use the course's brand color as their background/accent color.",
        "P1", ""))
    cases.append(c("2.2-008", "2.2", "Weekly Scheduler", "Add Session button creates session",
        "OWNER", "Classes exist",
        "1. Click the 'Add Session' button on the Schedule page\n2. Fill in the session form (Class, Date, Start Time, End Time, Room)\n3. Click Save/Create",
        "1. The new session appears on the calendar at the correct day and time\n2. Success toast appears",
        "P1", ""))

    # Story 2.3 — Conflict Detection
    cases.append(c("2.3-001", "2.3", "Conflict Detection", "Room double-booking shows warning",
        "OWNER", "Two different classes; one already has a session in Room A at 10 AM",
        "1. Create or move a session for a different class to Room A at 10 AM (same time and room)\n2. Observe the session form/dialog",
        "1. A non-blocking warning banner appears in the dialog showing: 'Conflict: Room A is booked by [Class X] at this time'\n2. The conflicting session details are shown",
        "P1", "Yes"))
    cases.append(c("2.3-002", "2.3", "Conflict Detection", "Teacher double-booking shows warning",
        "OWNER", "A teacher is assigned to Class A with a session at 2 PM; Class B has the same teacher",
        "1. Create a session for Class B at 2 PM (same teacher, overlapping time)\n2. Observe the form",
        "A warning banner appears: 'Conflict: [Teacher Name] is already teaching [Class A] at this time'.",
        "P1", "Yes"))
    cases.append(c("2.3-003", "2.3", "Conflict Detection", "Conflict icon on calendar blocks",
        "ALL", "A session with a conflict exists",
        "1. Navigate to the Schedule page\n2. Look at the session block that has a conflict",
        "The session block displays a warning icon (triangle with exclamation mark).",
        "P1", ""))
    cases.append(c("2.3-004", "2.3", "Conflict Detection", "Click conflict icon shows conflict details",
        "OWNER", "A session with a conflict exists on the calendar",
        "1. Click the warning icon on the conflicting session block",
        "A drawer or panel opens showing the specific conflicts: which room or teacher is double-booked, and with which other session.",
        "P1", ""))
    cases.append(c("2.3-005", "2.3", "Conflict Detection", "Suggestion for next available slot",
        "OWNER", "Creating a session that conflicts",
        "1. In the session creation dialog, trigger a conflict\n2. Look for suggestions",
        "The system suggests the next available time slot or an alternative room that doesn't have a conflict.",
        "P2", ""))
    cases.append(c("2.3-006", "2.3", "Conflict Detection", "Drag feedback: red for conflict, green for safe",
        "OWNER", "Sessions exist with some occupied time slots",
        "1. Start dragging a session block across different time slots",
        "1. Slots with conflicts highlight in red as you drag over them\n2. Available slots highlight in green",
        "P2", ""))
    cases.append(c("2.3-007", "2.3", "Conflict Detection", "Force Save override (Owner/Admin)",
        "OWNER", "A conflict exists when creating/moving a session",
        "1. Create/move a session that triggers a conflict warning\n2. Click 'Force Save' or 'Save Anyway'",
        "The session is saved despite the conflict. The conflict warning icon appears on the calendar block.",
        "P1", ""))

    # Story 2.4 — Attendance
    cases.append(c("2.4-001", "2.4", "Attendance", "Mark attendance for a past session",
        "TEACHER", "A past session exists for teacher's class; students are enrolled",
        "1. Navigate to the Schedule page\n2. Click on a past session block\n3. Click 'Mark Attendance' or open the attendance view\n4. Mark each student as Present, Absent, Late, or Excused using the toggle buttons\n5. Changes save automatically",
        "1. Session info header shows 'Attendance: [Course] - [Class] (Date)'\n2. Students are listed alphabetically\n3. Each student has 4 status buttons (Present, Absent, Late, Excused)\n4. Selections save immediately with visual confirmation",
        "P1", "Yes"))
    cases.append(c("2.4-002", "2.4", "Attendance", "Bulk Mark All Present",
        "TEACHER", "Attendance view is open for a session",
        "1. In the attendance view, click 'Mark All Present'\n2. Confirm in the dialog",
        "All students' statuses change to 'Present'. Each status button shows the Present state selected.",
        "P1", ""))
    cases.append(c("2.4-003", "2.4", "Attendance", "Bulk Mark All Absent",
        "TEACHER", "Attendance view is open",
        "1. Click 'Mark All Absent'\n2. Confirm",
        "All students' statuses change to 'Absent'.",
        "P2", ""))
    cases.append(c("2.4-004", "2.4", "Attendance", "Touch-friendly status buttons (44px+)",
        "TEACHER", "Attendance view on mobile",
        "1. Open attendance view on a mobile device or narrow screen\n2. Observe the status toggle buttons",
        "Each status button is at least 44x44 pixels, easy to tap without accidentally hitting adjacent buttons.",
        "P1", ""))
    cases.append(c("2.4-005", "2.4", "Attendance", "Cannot mark attendance for future sessions",
        "TEACHER", "A future session exists",
        "1. Click on a future session in the Schedule\n2. Look for attendance marking option",
        "The 'Mark Attendance' option is not available or is disabled for future sessions.",
        "P1", ""))
    cases.append(c("2.4-006", "2.4", "Attendance", "Teacher can only mark attendance for own classes",
        "TEACHER", "Teacher is assigned to Class A but not Class B; both have past sessions",
        "1. Navigate to Schedule\n2. Click on a past session for Class B (not teacher's class)\n3. Check for attendance option",
        "Attendance marking is not available for classes the teacher is not assigned to.",
        "P1", "Yes"))
    cases.append(c("2.4-007", "2.4", "Attendance", "Keyboard navigation in attendance view",
        "TEACHER", "Attendance view open on desktop",
        "1. Use Tab to move between students\n2. Use arrow keys to change status\n3. Press Enter to confirm",
        "Focus moves between students and status options using keyboard. Status can be changed without mouse.",
        "P3", ""))

    # Story 2.5 — Session CRUD
    cases.append(c("2.5-001", "2.5", "Session CRUD", "Create session by clicking empty time slot",
        "OWNER", "Classes exist",
        "1. Navigate to Schedule\n2. Click on an empty time slot in the calendar",
        "A session creation dialog opens with the date and time pre-filled based on the slot you clicked.",
        "P1", "Yes"))
    cases.append(c("2.5-002", "2.5", "Session CRUD", "Create session with all fields",
        "OWNER", "Classes and rooms exist",
        "1. Open the session creation dialog\n2. Select a Class\n3. Set Date, Start Time, End Time\n4. Select a Room\n5. Click Save",
        "1. Success toast appears\n2. Session appears on the calendar at the correct day/time\n3. Session block shows the class name and room",
        "P1", "Yes"))
    cases.append(c("2.5-003", "2.5", "Session CRUD", "Edit an existing session",
        "OWNER", "A session exists",
        "1. Click on a session block in the calendar\n2. Click Edit\n3. Change the room or time\n4. Click Save",
        "1. Success toast appears\n2. The session block updates on the calendar with the new time/room",
        "P1", ""))
    cases.append(c("2.5-004", "2.5", "Session CRUD", "Delete a single session",
        "OWNER", "A session exists",
        "1. Click on a session block\n2. Click Delete\n3. Choose 'Delete this session only'\n4. Confirm",
        "1. The session is removed from the calendar\n2. Other sessions in the same series (if recurring) remain",
        "P1", ""))
    cases.append(c("2.5-005", "2.5", "Session CRUD", "Delete all future sessions in a series",
        "OWNER", "A recurring session series exists",
        "1. Click on a session in a recurring series\n2. Click Delete\n3. Choose 'Delete all future sessions'\n4. Confirm",
        "1. The selected session and all future sessions in the series are removed\n2. Past sessions in the series remain",
        "P1", ""))
    cases.append(c("2.5-006", "2.5", "Session CRUD", "Create recurring weekly sessions",
        "OWNER", "A class exists",
        "1. Open session creation dialog\n2. Fill in class, day, time, room\n3. Set Recurrence to 'Weekly'\n4. Click Save",
        "Multiple session blocks appear on the calendar, one per week for 12 weeks, all on the same day and time.",
        "P1", "Yes"))
    cases.append(c("2.5-007", "2.5", "Session CRUD", "Create bi-weekly recurring sessions",
        "OWNER", "A class exists",
        "1. Create a session with Recurrence set to 'Bi-weekly'\n2. Save",
        "Session blocks appear every other week on the calendar (6 sessions over 12 weeks).",
        "P2", ""))
    cases.append(c("2.5-008", "2.5", "Session CRUD", "Drag-to-create session",
        "OWNER", "On the Schedule page",
        "1. Click and drag across a time range on an empty day column\n2. Release",
        "A session creation dialog opens with the start and end time pre-filled based on the drag range.",
        "P2", ""))
    cases.append(c("2.5-009", "2.5", "Session CRUD", "Room management in Settings",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > Rooms\n2. Click 'Add Room'\n3. Enter room name (e.g., 'Room 101')\n4. Save\n5. Verify the room appears in the list\n6. Edit the room name\n7. Delete a room",
        "1. New room appears in the rooms list\n2. Room name can be edited and saved\n3. Room can be deleted (removed from list)\n4. The room appears in session creation Room dropdown",
        "P1", ""))
    cases.append(c("2.5-010", "2.5", "Session CRUD", "Session notification when moved",
        "OWNER", "A session exists with enrolled students",
        "1. Drag a session to a new time slot\n2. Check the notification bell icon",
        "The notification bell shows an unread count. Students and teachers in the class should see a notification about the schedule change.",
        "P2", ""))

    return cases


# ──────────────────────────────────────────────
# EPIC 3 — Exercise Builder
# ──────────────────────────────────────────────
def epic3():
    cases = []
    # Story 3.1 — Core & Passage Management
    cases.append(c("3.1-001", "3.1", "Exercise Core", "Create a Reading exercise — available question types",
        "TEACHER", "Logged in as Teacher",
        "1. Navigate to Exercises page\n2. Click '+ New Exercise'\n3. Select 'Reading' as the skill\n4. Click the question type dropdown in a new section",
        "The dropdown shows Reading question types (R1 through R14): MCQ Single, MCQ Multi, True/False/Not Given, Yes/No/Not Given, Sentence Completion, Short Answer, Summary Word Bank, and Matching types, etc.",
        "P1", "Yes"))
    cases.append(c("3.1-002", "3.1", "Exercise Core", "Create a Listening exercise — available question types",
        "TEACHER", "Logged in",
        "1. Navigate to Exercises > + New Exercise\n2. Select 'Listening' as the skill\n3. Click the question type dropdown",
        "The dropdown shows Listening question types (L1 through L6): Form/Note/Table, MCQ, Matching, Map/Plan, Sentence Completion, Short Answer.",
        "P1", ""))
    cases.append(c("3.1-003", "3.1", "Exercise Core", "Create a Writing exercise — available question types",
        "TEACHER", "Logged in",
        "1. Navigate to Exercises > + New Exercise\n2. Select 'Writing' as the skill\n3. Check available question types",
        "Writing question types are available (W1 through W3): Task 1 Academic, Task 1 General, Task 2 Essay.",
        "P1", ""))
    cases.append(c("3.1-004", "3.1", "Exercise Core", "Create a Speaking exercise — available question types",
        "TEACHER", "Logged in",
        "1. Navigate to Exercises > + New Exercise\n2. Select 'Speaking' as the skill\n3. Check available types",
        "Speaking question types are available (S1 through S3): Part 1 QA, Part 2 Cue Card, Part 3 Discussion.",
        "P1", ""))
    cases.append(c("3.1-005", "3.1", "Passage Editor", "Add passage with auto paragraph lettering",
        "TEACHER", "Creating a Reading exercise",
        "1. In the exercise editor with Reading skill selected\n2. Enter passage text with multiple paragraphs (separated by blank lines)\n3. Observe the paragraph labels",
        "Each paragraph is automatically labeled A, B, C, etc. Labels appear visually next to each paragraph.",
        "P1", ""))
    cases.append(c("3.1-006", "3.1", "Exercise Core", "Enter exercise title and instructions",
        "TEACHER", "Exercise editor is open",
        "1. Enter a title in the Title field\n2. Enter instructions in the Instructions field\n3. Click Save Draft",
        "1. Title and instructions are saved\n2. Success toast appears\n3. Returning to the exercise later shows the saved title and instructions",
        "P1", ""))
    cases.append(c("3.1-007", "3.1", "Exercise Core", "Auto-save triggers every 30 seconds",
        "TEACHER", "Exercise editor is open with unsaved changes",
        "1. Make a change (e.g., edit the title)\n2. Wait 30 seconds without clicking Save\n3. Observe the save indicator",
        "An auto-save indicator appears (e.g., 'Saving...' then 'Saved'). Closing and reopening the exercise shows the changes were saved.",
        "P2", ""))
    cases.append(c("3.1-008", "3.1", "Exercise Core", "Preview mode shows student view",
        "TEACHER", "Exercise with questions created",
        "1. In the exercise editor, click 'Preview'\n2. Observe the preview",
        "The exercise displays exactly as a student would see it: passage text, questions with input fields, navigation between questions. Editor controls are hidden.",
        "P1", "Yes"))
    cases.append(c("3.1-009", "3.1", "Exercise Core", "Save as Draft — exercise not assignable",
        "TEACHER", "A Draft exercise exists",
        "1. Go to Assignments page\n2. Click '+ New Assignment'\n3. Look for the Draft exercise in the exercise selection",
        "The Draft exercise does not appear in the exercise selection list (only Published exercises are available for assignment).",
        "P1", "Yes"))
    cases.append(c("3.1-010", "3.1", "Exercise Core", "Publish exercise — becomes assignable",
        "TEACHER", "A Draft exercise with questions exists",
        "1. Open the exercise editor\n2. Click 'Publish'\n3. Go to Assignments > + New Assignment\n4. Look for the exercise in the selection",
        "1. The exercise status changes to 'Published' (green badge)\n2. The exercise now appears in the assignment exercise selection list",
        "P1", "Yes"))

    # MCQ and question types
    cases.append(c("3.1-011", "3.1", "R1 - MCQ Single", "Add MCQ Single Choice question",
        "TEACHER", "Reading exercise open in editor",
        "1. Add a new section with type 'MCQ Single'\n2. Click '+ Add Question'\n3. Enter question text\n4. Enter 4 options (A, B, C, D)\n5. Mark option B as the correct answer\n6. Save",
        "1. Question appears in the section with question text and 4 options\n2. The correct answer indicator shows on option B\n3. In Preview mode, options appear as selectable buttons",
        "P1", "Yes"))
    cases.append(c("3.1-012", "3.1", "R2 - MCQ Multi", "Add MCQ Multiple Choice question",
        "TEACHER", "Reading exercise open in editor",
        "1. Add a section with type 'MCQ Multi'\n2. Add a question\n3. Enter 5 options\n4. Mark 2 options as correct\n5. Set max selections\n6. Save",
        "1. Question shows with multiple correct answers marked\n2. In Preview, student can select multiple options up to the max",
        "P1", ""))
    cases.append(c("3.1-013", "3.1", "R3 - TFNG", "Add True/False/Not Given question",
        "TEACHER", "Reading exercise open",
        "1. Add a section with type 'True/False/Not Given'\n2. Add a question/statement\n3. Set the answer key (True, False, or Not Given)\n4. Save",
        "1. Question appears with the correct answer set\n2. In Preview, student sees three options: True, False, Not Given",
        "P1", ""))
    cases.append(c("3.1-014", "3.1", "R4 - YNNG", "Add Yes/No/Not Given question",
        "TEACHER", "Reading exercise open",
        "1. Add a section with type 'Yes/No/Not Given'\n2. Add a question\n3. Set the answer\n4. Save",
        "1. Answer key is saved\n2. In Preview, student sees: Yes, No, Not Given options",
        "P1", ""))
    cases.append(c("3.1-015", "3.1", "R5 - Sentence Completion", "Add Sentence Completion question",
        "TEACHER", "Reading exercise open",
        "1. Add a Sentence Completion section\n2. Add a question with a blank marker in the sentence\n3. Set the answer text\n4. Save",
        "1. Question shows with a blank/gap in the sentence\n2. In Preview, student sees a text input at the blank position",
        "P1", ""))
    cases.append(c("3.1-016", "3.1", "R6 - Short Answer", "Add Short Answer question with word limit",
        "TEACHER", "Reading exercise open",
        "1. Add a Short Answer section\n2. Add a question\n3. Set word limit (e.g., 'No more than 3 words')\n4. Set the answer\n5. Save",
        "1. Question shows with the answer and word limit\n2. In Preview, student sees a text input with word limit indicator",
        "P1", ""))
    cases.append(c("3.1-017", "3.1", "R7 - Word Bank", "Add Summary with Word Bank question",
        "TEACHER", "Reading exercise open",
        "1. Add a Summary Word Bank section\n2. Enter summary text with blanks\n3. Add word bank items (correct answers + distractors)\n4. Save",
        "1. Summary text appears with blank markers\n2. Word bank list shows below\n3. In Preview, student selects words from dropdown to fill blanks",
        "P1", ""))
    cases.append(c("3.1-018", "3.1", "Matching Types", "Add Matching Headings question (R9-R12)",
        "TEACHER", "Reading exercise open",
        "1. Add a Matching section\n2. Add left column items (e.g., paragraph headings)\n3. Add right column items (e.g., paragraphs A-F)\n4. Set the correct matches\n5. Save",
        "1. Two columns of items appear with match connections\n2. In Preview, student sees dropdown selections to match items",
        "P1", ""))
    cases.append(c("3.1-019", "3.1", "R13 - Note/Table", "Add Note/Table/Flowchart completion question",
        "TEACHER", "Reading exercise open",
        "1. Add a Note/Table/Flowchart section\n2. Add structured fields with blanks\n3. Set answers for each blank\n4. Save",
        "1. Structured form layout appears with blank fields\n2. In Preview, student fills in text inputs within the structure",
        "P1", ""))
    cases.append(c("3.1-020", "3.1", "R14 - Diagram", "Add Diagram Labelling question",
        "TEACHER", "Reading exercise open",
        "1. Add a Diagram Labelling section\n2. Upload an image\n3. Place label markers on the diagram\n4. Set answers for each label\n5. Save",
        "1. Image displays with label markers positioned on it\n2. In Preview, student sees the diagram with text input fields at each label position",
        "P2", ""))
    cases.append(c("3.1-021", "3.1", "Question Management", "Drag-to-reorder questions within a section",
        "TEACHER", "A section with 3+ questions",
        "1. In the exercise editor, find a section with multiple questions\n2. Click and drag a question to a different position\n3. Release",
        "The question moves to the new position. Question numbering updates automatically.",
        "P1", ""))
    cases.append(c("3.1-022", "3.1", "Question Management", "Delete a question",
        "TEACHER", "A section with questions",
        "1. Click the delete icon on a question\n2. Confirm",
        "The question is removed. Remaining questions renumber automatically.",
        "P1", ""))

    # Listening specific
    cases.append(c("3.1-023", "3.1", "Listening - Audio", "Upload audio file for Listening exercise",
        "TEACHER", "Creating a Listening exercise",
        "1. Select 'Listening' as the skill\n2. Click the audio upload area\n3. Select an audio file (MP3/WAV)\n4. Wait for upload",
        "1. Audio file uploads with a progress indicator\n2. An audio player appears showing the audio waveform or controls\n3. Play/Pause/Seek controls are functional",
        "P1", "Yes"))
    cases.append(c("3.1-024", "3.1", "Listening - Audio", "Add audio section markers",
        "TEACHER", "Listening exercise with audio uploaded",
        "1. In the audio settings, add section break markers at specific timestamps\n2. Label each section",
        "Section markers appear on the audio timeline. Each section can have its own set of questions.",
        "P1", ""))
    cases.append(c("3.1-025", "3.1", "Listening - Audio", "Toggle transcript display",
        "TEACHER", "Listening exercise with audio",
        "1. Toggle 'Show Transcript' on\n2. Preview the exercise",
        "In Preview mode, a transcript panel appears alongside the audio player.",
        "P2", ""))

    # Writing specific
    cases.append(c("3.1-026", "3.1", "Writing Task", "Set up Writing Task 1 with stimulus image",
        "TEACHER", "Creating a Writing exercise",
        "1. Select 'Writing' skill\n2. Select type 'Task 1 Academic'\n3. Enter the prompt/task description\n4. Upload a stimulus image (graph/chart)\n5. Set minimum word count\n6. Save",
        "1. Task prompt displays with the uploaded image\n2. Word count minimum is saved\n3. In Preview, student sees the prompt, image, and a writing area with live word count",
        "P1", ""))
    cases.append(c("3.1-027", "3.1", "Writing Task", "Set letter tone for General Writing Task",
        "TEACHER", "Creating a Writing exercise with W1 General type",
        "1. Select type 'Task 1 General'\n2. Set the letter tone (Formal, Informal, or Semi-formal)\n3. Save",
        "The tone selector is saved and visible in the exercise settings.",
        "P2", ""))

    # Speaking specific
    cases.append(c("3.1-028", "3.1", "Speaking Task", "Set up Speaking Part 2 with cue card",
        "TEACHER", "Creating a Speaking exercise",
        "1. Select 'Speaking' skill\n2. Select type 'Part 2 Cue Card'\n3. Enter the cue card text (topic + bullet points)\n4. Set prep time (e.g., 60 seconds)\n5. Set speaking time (e.g., 120 seconds)\n6. Save",
        "1. Cue card text displays in the editor\n2. Prep and speaking times are saved\n3. In Preview, student sees the cue card with a countdown timer",
        "P1", ""))
    cases.append(c("3.1-029", "3.1", "Speaking Task", "Set max recording duration",
        "TEACHER", "Creating a Speaking exercise",
        "1. Set the max recording duration (e.g., 180 seconds)\n2. Save",
        "Recording duration limit is saved. In Preview, the recording stops automatically when the limit is reached.",
        "P2", ""))

    # Exercise list management
    cases.append(c("3.1-030", "3.1", "Exercise List", "View exercises table with filters",
        "TEACHER", "Multiple exercises exist",
        "1. Navigate to the Exercises page\n2. Observe the table and filter options",
        "1. Table shows: Status badge, Title, Skill icon, Tags, Question count, Assignment count, Actions\n2. Filter dropdowns are available for: Skill, Status, Tags\n3. A search field is available\n4. Pagination controls appear at the bottom",
        "P1", ""))
    cases.append(c("3.1-031", "3.1", "Exercise List", "Filter exercises by skill",
        "TEACHER", "Exercises exist with different skills",
        "1. On Exercises page, select 'Reading' from the Skill filter",
        "Only Reading exercises are shown. Other skills are hidden.",
        "P1", ""))
    cases.append(c("3.1-032", "3.1", "Exercise List", "Filter exercises by status",
        "TEACHER", "Exercises exist with Draft, Published, and Archived statuses",
        "1. Select 'Draft' from the Status filter",
        "Only Draft exercises are shown.",
        "P1", ""))
    cases.append(c("3.1-033", "3.1", "Exercise List", "Search exercises by title",
        "TEACHER", "Multiple exercises exist",
        "1. Type a keyword in the search field",
        "Table filters to show only exercises whose title contains the search keyword.",
        "P1", ""))
    cases.append(c("3.1-034", "3.1", "Exercise List", "Copy/duplicate an exercise",
        "TEACHER", "An exercise exists",
        "1. Click the Copy action on an exercise",
        "1. A new exercise is created with the same content and a modified title (e.g., 'Copy of [title]')\n2. The copy appears in the exercise list as Draft",
        "P1", ""))
    cases.append(c("3.1-035", "3.1", "Exercise List", "Archive an exercise",
        "TEACHER", "A Published exercise exists (no active assignments)",
        "1. Click the Archive action on a Published exercise\n2. Confirm",
        "The exercise's status changes to 'Archived' (muted badge). It no longer appears in the Published filter.",
        "P1", ""))
    cases.append(c("3.1-036", "3.1", "Exercise List", "Delete an exercise",
        "TEACHER", "A Draft or Archived exercise exists",
        "1. Click the Delete (trash) action on an exercise\n2. Confirm",
        "1. Confirmation dialog appears\n2. The exercise is removed from the list\n3. Success toast appears",
        "P1", ""))
    cases.append(c("3.1-037", "3.1", "Exercise List", "Create assignment from exercise action",
        "TEACHER", "A Published exercise exists",
        "1. Click the 'Create Assignment' action on a Published exercise",
        "The Create Assignment dialog opens with the exercise pre-selected.",
        "P1", ""))
    cases.append(c("3.1-038", "3.1", "Exercise List", "Toggle table/grid view",
        "TEACHER", "On the Exercises page",
        "1. Click the view toggle (table/grid icon)",
        "The exercise list switches between table view and grid/card view. All exercises remain visible.",
        "P2", ""))
    cases.append(c("3.1-039", "3.1", "Exercise List", "Filter by tags",
        "TEACHER", "Exercises exist with tags assigned",
        "1. Select a tag from the Tags filter",
        "Only exercises with that tag are shown.",
        "P2", ""))
    cases.append(c("3.1-040", "3.1", "Exercise List", "Pagination with 20 items per page",
        "TEACHER", "More than 20 exercises exist",
        "1. Navigate to Exercises page\n2. Count items on the first page\n3. Click page 2",
        "1. First page shows 20 exercises\n2. Page 2 shows the next batch\n3. Pagination controls indicate total pages",
        "P2", ""))

    # Tags management
    cases.append(c("3.1-041", "3.1", "Tags", "Create, edit, and delete tags",
        "OWNER", "Logged in as Owner",
        "1. Navigate to Settings > Tags\n2. Click 'Add Tag' and enter a tag name (e.g., 'Vocabulary')\n3. Save\n4. Edit the tag name\n5. Delete a tag",
        "1. New tag appears in the tag list\n2. Tag name can be updated\n3. Tag can be deleted\n4. Tags appear as filter options on the Exercises page",
        "P1", ""))

    # Assignments
    cases.append(c("3.2-001", "3.2", "Assignments", "Create assignment for a class",
        "TEACHER", "A Published exercise and a class with students exist",
        "1. Navigate to Assignments page\n2. Click '+ New Assignment'\n3. Select a Published exercise\n4. Choose 'Assign to Classes'\n5. Select a class\n6. Set a due date\n7. Optionally set a time limit (minutes)\n8. Add instructions\n9. Click Create",
        "1. Success toast appears\n2. The assignment appears in the Assignments table with status 'Open'\n3. Students in the selected class see the assignment on their dashboard",
        "P1", "Yes"))
    cases.append(c("3.2-002", "3.2", "Assignments", "Create assignment for individual students",
        "TEACHER", "A Published exercise and students exist",
        "1. Click '+ New Assignment'\n2. Select exercise\n3. Choose 'Assign to Students'\n4. Select individual students\n5. Set due date and time limit\n6. Create",
        "1. Assignment is created\n2. Only the selected students see the assignment on their dashboard",
        "P1", ""))
    cases.append(c("3.2-003", "3.2", "Assignments", "View assignments table with filters",
        "TEACHER", "Multiple assignments exist",
        "1. Navigate to Assignments page\n2. Observe the table and filter options",
        "Table shows: Status badge, Title, Exercise, Skill icon, Class, Due date, Time limit, Questions, Submissions, Actions. Filters: Status, Skill, Class, Due date range. Search bar available.",
        "P1", ""))
    cases.append(c("3.2-004", "3.2", "Assignments", "Filter assignments by status",
        "TEACHER", "Assignments with different statuses exist",
        "1. Select 'Open' from the Status filter",
        "Only Open assignments are shown.",
        "P1", ""))
    cases.append(c("3.2-005", "3.2", "Assignments", "Close an assignment",
        "TEACHER", "An Open assignment exists",
        "1. Click the Close action on an assignment\n2. Confirm",
        "Assignment status changes to 'Closed'. Students can no longer submit.",
        "P1", "Yes"))
    cases.append(c("3.2-006", "3.2", "Assignments", "Reopen a closed assignment",
        "TEACHER", "A Closed assignment exists",
        "1. Click Reopen on the closed assignment\n2. Set a new due date\n3. Confirm",
        "Assignment status returns to 'Open' with the new due date. Students can submit again.",
        "P1", ""))
    cases.append(c("3.2-007", "3.2", "Assignments", "Archive an assignment",
        "TEACHER", "A Closed assignment exists",
        "1. Click Archive on the assignment\n2. Confirm",
        "Assignment status changes to 'Archived'. It no longer appears in the default list (visible only when filtering by Archived).",
        "P2", ""))
    cases.append(c("3.2-008", "3.2", "Assignments", "Edit assignment due date and time limit",
        "TEACHER", "An Open assignment exists",
        "1. Click Edit (pencil) on an assignment\n2. Change the due date\n3. Change the time limit\n4. Save",
        "1. Success toast appears\n2. The updated due date and time limit appear in the table",
        "P1", ""))
    cases.append(c("3.2-009", "3.2", "Assignments", "Delete an assignment",
        "TEACHER", "An assignment exists (preferably with no submissions)",
        "1. Click Delete (trash) on an assignment\n2. Confirm",
        "1. Confirmation dialog appears\n2. Assignment is removed from the table\n3. Students no longer see it on their dashboard",
        "P1", ""))
    cases.append(c("3.2-010", "3.2", "Assignments", "Submission count reflects completed submissions",
        "TEACHER", "An assignment exists; students have submitted",
        "1. Look at the Submissions column in the Assignments table",
        "The count shows the number of completed submissions out of total assigned students (e.g., '3/15').",
        "P1", ""))

    # Mock Tests
    cases.append(c("3.3-001", "3.3", "Mock Tests", "Create a new mock test",
        "TEACHER", "Logged in",
        "1. Navigate to Mock Tests page\n2. Click '+ New Mock Test'\n3. Enter title\n4. Select type (Academic or General Training)\n5. Add description\n6. Click Create",
        "1. Mock test is created\n2. You are taken to the mock test editor\n3. The mock test appears in the Mock Tests table",
        "P1", "Yes"))
    cases.append(c("3.3-002", "3.3", "Mock Tests", "Add exercises to mock test sections",
        "TEACHER", "A mock test exists; Published exercises exist",
        "1. Open the mock test editor\n2. Add exercises to each section (Reading, Listening, Writing, Speaking)\n3. Save",
        "Each section shows the added exercises with their question counts.",
        "P1", ""))
    cases.append(c("3.3-003", "3.3", "Mock Tests", "Filter mock tests by status",
        "TEACHER", "Mock tests with different statuses exist",
        "1. On Mock Tests page, select a status filter",
        "Table filters to show only mock tests with the selected status.",
        "P2", ""))
    cases.append(c("3.3-004", "3.3", "Mock Tests", "Filter mock tests by type",
        "TEACHER", "Both Academic and General Training mock tests exist",
        "1. Select 'Academic' from the Type filter",
        "Only Academic mock tests are shown.",
        "P2", ""))
    cases.append(c("3.3-005", "3.3", "Mock Tests", "Publish a mock test",
        "TEACHER", "A draft mock test with exercises in each section",
        "1. Click Publish on the mock test\n2. Confirm",
        "Status changes to 'Published'. The mock test can now be assigned.",
        "P1", ""))
    cases.append(c("3.3-006", "3.3", "Mock Tests", "Archive a mock test",
        "TEACHER", "A published mock test",
        "1. Click Archive\n2. Confirm",
        "Status changes to 'Archived'. Mock test is no longer assignable.",
        "P2", ""))
    cases.append(c("3.3-007", "3.3", "Mock Tests", "Delete a mock test",
        "TEACHER", "A draft or archived mock test",
        "1. Click Delete\n2. Confirm",
        "Mock test is removed from the table.",
        "P1", ""))

    # AI Generation
    cases.append(c("3.4-001", "3.4", "AI Generation", "Generate exercise content via AI",
        "TEACHER", "Exercise editor open; AI panel visible",
        "1. In the exercise editor, open the AI generation panel\n2. Upload a document or describe the desired content\n3. Click Generate\n4. Wait for AI to produce content",
        "1. Loading indicator appears during generation\n2. Generated questions/content appears in the editor\n3. Content can be reviewed and edited before saving",
        "P1", ""))

    # Timer
    cases.append(c("3.5-001", "3.5", "Timer Settings", "Configure timer for an exercise",
        "TEACHER", "Exercise editor open",
        "1. In the exercise editor, find the Timer settings panel\n2. Set a time limit (e.g., 30 minutes)\n3. Save",
        "Timer setting is saved. When the exercise is assigned with this timer, students see a countdown during submission.",
        "P2", ""))

    return cases


# ──────────────────────────────────────────────
# EPIC 4 — Submissions
# ──────────────────────────────────────────────
def epic4():
    cases = []
    cases.append(c("4.1-001", "4.1", "Submission UI", "Start a submission for an assigned assignment",
        "STUDENT", "Logged in as Student; an Open assignment is visible on dashboard",
        "1. Navigate to the Dashboard\n2. Find the assignment card\n3. Click 'Start' on the assignment card",
        "1. Full-screen submission view opens (navigation sidebar hidden)\n2. The exercise title appears in the header\n3. The first question is displayed\n4. Question navigation pills (1, 2, 3...) appear at the top",
        "P1", "Yes"))
    cases.append(c("4.1-002", "4.1", "Submission UI", "Navigate between questions",
        "STUDENT", "In the submission view; exercise has 5+ questions",
        "1. Click the 'Next' button\n2. Click the 'Previous' button\n3. Click question pill #3",
        "1. Next: advances to the next question\n2. Previous: goes back to the prior question\n3. Clicking pill #3: jumps directly to question 3",
        "P1", "Yes"))
    cases.append(c("4.1-003", "4.1", "Submission UI", "Question pills show answered status",
        "STUDENT", "In the submission view",
        "1. Answer question 1 and question 3\n2. Leave question 2 unanswered\n3. Observe the question pills",
        "Pills for questions 1 and 3 are visually filled/highlighted (indicating answered). Pill for question 2 remains unfilled/empty.",
        "P1", ""))
    cases.append(c("4.1-004", "4.1", "MCQ Interaction", "MCQ renders as tap-friendly buttons (44px+)",
        "STUDENT", "Viewing an MCQ question on mobile",
        "1. Open an MCQ question in the submission view on a mobile device\n2. Observe the option buttons",
        "Each option button is at least 44px tall, easy to tap. Selecting an option highlights it clearly.",
        "P1", ""))
    cases.append(c("4.1-005", "4.1", "MCQ Interaction", "Single-select MCQ allows only one answer",
        "STUDENT", "Viewing an MCQ Single question",
        "1. Tap option A\n2. Tap option C",
        "1. Tapping A selects it\n2. Tapping C deselects A and selects C (only one option selected at a time)",
        "P1", ""))
    cases.append(c("4.1-006", "4.1", "MCQ Interaction", "Multi-select MCQ allows multiple answers",
        "STUDENT", "Viewing an MCQ Multi question",
        "1. Tap option A\n2. Tap option C\n3. Tap option D",
        "All three options (A, C, D) are selected simultaneously. The max selection limit (if set) is enforced.",
        "P1", ""))
    cases.append(c("4.1-007", "4.1", "Fill-in-Blank", "Fill-in-blank shows expandable text input",
        "STUDENT", "Viewing a Sentence Completion question",
        "1. Tap the blank/gap in the sentence\n2. Start typing",
        "A text input appears at the blank position. Word limit indicator is shown if applicable.",
        "P1", ""))
    cases.append(c("4.1-008", "4.1", "Word Bank", "Word bank question shows dropdown",
        "STUDENT", "Viewing a Summary Word Bank question",
        "1. Tap a blank in the summary\n2. Observe the dropdown",
        "A dropdown or tap-to-select menu appears with the word bank items. Selecting a word fills the blank.",
        "P1", ""))
    cases.append(c("4.1-009", "4.1", "Matching", "Matching question uses dropdowns",
        "STUDENT", "Viewing a Matching question",
        "1. Look at the matching items\n2. Tap the dropdown next to an item\n3. Select a match",
        "Each item has a dropdown for selecting the match. Selected matches are clearly displayed.",
        "P1", ""))
    cases.append(c("4.1-010", "4.1", "TFNG/YNNG", "True/False/Not Given shows three options",
        "STUDENT", "Viewing a TFNG statement",
        "1. Read the statement\n2. Tap one of the three options",
        "Three buttons appear: True, False, Not Given. Selecting one highlights it and deselects others.",
        "P1", ""))
    cases.append(c("4.1-011", "4.1", "Reading Passage", "Reading passage panel alongside questions",
        "STUDENT", "Viewing a Reading exercise in submission view",
        "1. Observe the layout",
        "The reading passage appears in a side panel (right on desktop, toggleable on mobile). Paragraphs are labeled A, B, C. The passage is scrollable independently of the questions.",
        "P1", "Yes"))
    cases.append(c("4.1-012", "4.1", "Listening Audio", "Audio player with section markers",
        "STUDENT", "Viewing a Listening exercise in submission view",
        "1. Observe the audio player\n2. Press Play\n3. Notice section markers",
        "1. Audio player appears with Play/Pause/Seek controls\n2. Audio plays correctly\n3. Section markers are visible on the timeline",
        "P1", "Yes"))
    cases.append(c("4.1-013", "4.1", "Writing Submission", "Writing area with live word count",
        "STUDENT", "Viewing a Writing task in submission view",
        "1. Start typing in the writing area\n2. Observe the word count",
        "1. Text area is available for writing\n2. A live word count updates as you type\n3. Minimum word count requirement is shown if set",
        "P1", ""))
    cases.append(c("4.1-014", "4.1", "Auto-save", "Answers auto-save periodically",
        "STUDENT", "In the submission view; have answered some questions",
        "1. Answer 2-3 questions\n2. Observe the save indicator in the header\n3. Close the browser tab\n4. Navigate back to the assignment and click 'Continue'",
        "1. Save indicator shows 'Saving...' then 'Saved'\n2. After reopening, your previous answers are restored",
        "P1", "Yes"))
    cases.append(c("4.1-015", "4.1", "Auto-save", "Offline save indicator",
        "STUDENT", "In the submission view",
        "1. Disconnect from the internet (airplane mode or disable WiFi)\n2. Continue answering questions\n3. Observe the save indicator",
        "1. Save indicator changes to 'Offline' or shows an offline icon\n2. Answers are still saved locally\n3. When you reconnect, answers sync and indicator returns to 'Saved'",
        "P1", "Yes"))
    cases.append(c("4.1-016", "4.1", "Submit", "Submit completed exercise",
        "STUDENT", "All questions answered in the submission view",
        "1. Click the 'Submit' button\n2. A confirmation dialog appears\n3. Click 'Confirm'",
        "1. Confirmation dialog shows a summary (e.g., 'Submit X answers?')\n2. After confirming, a 'Thank you' or completion page appears\n3. You cannot go back and change answers\n4. Assignment status changes on the dashboard",
        "P1", "Yes"))
    cases.append(c("4.1-017", "4.1", "Submit", "Submit with unanswered questions shows warning",
        "STUDENT", "Some questions left unanswered",
        "1. Click 'Submit' with unanswered questions\n2. Observe the confirmation dialog",
        "The confirmation dialog warns about unanswered questions (e.g., 'You have 3 unanswered questions. Are you sure you want to submit?').",
        "P1", ""))
    cases.append(c("4.1-018", "4.1", "Timer", "Timer countdown during timed submission",
        "STUDENT", "Assignment has a time limit set",
        "1. Start the submission\n2. Observe the timer in the header",
        "1. A countdown timer appears showing remaining time (e.g., '29:45')\n2. Timer counts down in real-time\n3. When time reaches 0, submission is automatically submitted",
        "P1", "Yes"))
    cases.append(c("4.1-019", "4.1", "Timer", "Timer warning near expiry",
        "STUDENT", "Timed submission with ~5 minutes remaining",
        "1. Let the timer count down to under 5 minutes\n2. Observe the timer display",
        "The timer changes color or style to indicate urgency (e.g., turns red). A warning may appear.",
        "P2", ""))
    cases.append(c("4.1-020", "4.1", "Mobile Layout", "Submission view responsive on mobile",
        "STUDENT", "On a mobile device (<768px)",
        "1. Open a submission on mobile\n2. Navigate between questions\n3. Answer questions",
        "1. Layout is single-column and usable on small screens\n2. Reading passage is in a toggleable panel (not always visible)\n3. All controls are touch-friendly (44px+ tap targets)\n4. Question pills scroll horizontally if many questions",
        "P1", "Yes"))
    cases.append(c("4.1-021", "4.1", "Submission Access", "Student can only see assignments assigned to them",
        "STUDENT", "Student is in Class A but not Class B; assignments exist for both classes",
        "1. Log in as the Student\n2. Check the Dashboard for assignments",
        "Only assignments for Class A appear. Assignments for Class B are not visible.",
        "P1", "Yes"))
    cases.append(c("4.1-022", "4.1", "Submission Access", "Student cannot take a closed assignment",
        "STUDENT", "An assignment has been closed by the teacher",
        "1. Look for the closed assignment on the Dashboard",
        "The closed assignment either does not appear or shows as 'Closed' with no 'Start' button.",
        "P1", ""))
    cases.append(c("4.1-023", "4.1", "Submission Access", "Overdue assignment is visually highlighted",
        "STUDENT", "An assignment with a past due date exists",
        "1. Look at the Dashboard\n2. Find the overdue assignment",
        "The overdue assignment appears in the 'Overdue' group with a red/urgent visual indicator (e.g., red badge or text color).",
        "P1", ""))
    cases.append(c("4.1-024", "4.1", "Skill Filter", "Filter assignments by skill on student dashboard",
        "STUDENT", "Assignments exist for different skills",
        "1. On the Dashboard, use the Skill filter dropdown\n2. Select 'Reading'",
        "Only Reading assignments are shown. Other skill assignments are hidden.",
        "P2", ""))

    return cases


# ──────────────────────────────────────────────
# EPIC 5 — AI Grading
# ──────────────────────────────────────────────
def epic5():
    cases = []
    # Story 5.1 — AI Analysis
    cases.append(c("5.1-001", "5.1", "AI Analysis", "Writing submission shows AI analysis in grading",
        "TEACHER", "A student has submitted a Writing exercise",
        "1. Navigate to the Grading page\n2. Click on the Writing submission from the queue",
        "1. The AI analysis panel shows band scores per IELTS criteria (Task Achievement, Coherence & Cohesion, Lexical Resource, Grammatical Range & Accuracy)\n2. Individual feedback items with grammar/vocabulary highlights are displayed\n3. An overall band score suggestion is shown",
        "P1", "Yes"))
    cases.append(c("5.1-002", "5.1", "AI Analysis", "Speaking submission shows AI analysis in grading",
        "TEACHER", "A student has submitted a Speaking exercise",
        "1. Navigate to Grading\n2. Click on the Speaking submission",
        "AI analysis shows band scores for Speaking criteria (Fluency & Coherence, Lexical Resource, Grammatical Range & Accuracy, Pronunciation). Feedback items are listed.",
        "P1", ""))
    cases.append(c("5.1-003", "5.1", "AI Analysis", "Reading submission is auto-graded (no AI panel)",
        "TEACHER", "A student has submitted a Reading exercise",
        "1. Navigate to Grading\n2. Click on the Reading submission",
        "The submission shows auto-graded results with correct/incorrect marks per question. No AI analysis panel or band scores are shown (objective marking only).",
        "P1", "Yes"))
    cases.append(c("5.1-004", "5.1", "AI Analysis", "Listening submission is auto-graded (no AI panel)",
        "TEACHER", "A student has submitted a Listening exercise",
        "1. Navigate to Grading\n2. Click on the Listening submission",
        "Same as Reading: auto-graded with correct/incorrect marks. No AI analysis panel.",
        "P1", ""))
    cases.append(c("5.1-005", "5.1", "AI Analysis", "AI analysis loading state",
        "TEACHER", "A student just submitted a Writing exercise (AI may still be processing)",
        "1. Navigate to Grading immediately after submission\n2. Open the submission",
        "If AI is still processing, a loading skeleton or 'Analyzing...' indicator appears in the feedback panel. Once complete, the AI results populate.",
        "P1", ""))
    cases.append(c("5.1-006", "5.1", "AI Analysis", "AI analysis failure shows error and manual grading option",
        "TEACHER", "AI analysis has failed for a submission",
        "1. Open a submission where AI analysis failed",
        "1. An error message appears in the feedback panel (e.g., 'AI analysis failed')\n2. The failure reason is shown\n3. A 'Re-analyze' button is available\n4. Teacher can still grade manually by entering scores directly",
        "P1", ""))
    cases.append(c("5.1-007", "5.1", "AI Analysis", "Re-trigger AI analysis",
        "TEACHER", "AI analysis has failed or completed",
        "1. Click 'Re-analyze' button on a submission",
        "1. Loading indicator appears\n2. New AI results replace the old ones (or populate if previously failed)",
        "P2", ""))

    # Story 5.2 — Split-Screen Grading
    cases.append(c("5.2-001", "5.2", "Grading Workbench", "Split-screen layout: student work left, AI feedback right",
        "TEACHER", "A Writing submission with AI analysis exists",
        "1. Open the submission from the Grading queue",
        "1. Left pane shows the student's essay/written answer with the question prompt\n2. Right pane shows AI feedback cards with band scores and feedback items\n3. Panes are resizable (drag the divider)",
        "P1", "Yes"))
    cases.append(c("5.2-002", "5.2", "Grading Workbench", "Navigate between submissions",
        "TEACHER", "Multiple submissions in the grading queue",
        "1. Open a submission\n2. Click the 'Next' arrow\n3. Click the 'Previous' arrow",
        "1. Next submission loads with minimal delay (<500ms)\n2. Previous submission loads\n3. Position indicator shows 'X of Y' (e.g., '3 of 15')",
        "P1", "Yes"))
    cases.append(c("5.2-003", "5.2", "Grading Workbench", "Keyboard shortcuts for navigation",
        "TEACHER", "Grading workbench open on desktop",
        "1. Press the Right Arrow key\n2. Press the Left Arrow key",
        "1. Right arrow navigates to the next submission\n2. Left arrow navigates to the previous submission",
        "P2", ""))
    cases.append(c("5.2-004", "5.2", "Grading Workbench", "Approve individual AI feedback items",
        "TEACHER", "Grading workbench open; AI feedback items visible",
        "1. Click 'Approve' (checkmark) on a specific feedback item",
        "The feedback item is marked as approved (visual indicator changes, e.g., green check).",
        "P1", "Yes"))
    cases.append(c("5.2-005", "5.2", "Grading Workbench", "Reject individual AI feedback items",
        "TEACHER", "Grading workbench open; AI feedback items visible",
        "1. Click 'Reject' (X) on a specific feedback item",
        "The feedback item is marked as rejected (crossed out or dimmed).",
        "P1", ""))
    cases.append(c("5.2-006", "5.2", "Grading Workbench", "Bulk approve all AI feedback items",
        "TEACHER", "Multiple unapproved feedback items",
        "1. Click 'Approve All' button",
        "All feedback items are marked as approved simultaneously.",
        "P1", ""))
    cases.append(c("5.2-007", "5.2", "Grading Workbench", "Override AI band score",
        "TEACHER", "AI has suggested band scores",
        "1. Click on a band score (e.g., Task Achievement: 6.5)\n2. Change it to a different score (e.g., 7.0)\n3. Save/Apply",
        "1. The score updates to the teacher's value\n2. The overall score recalculates\n3. The override is visually distinct from the AI suggestion (e.g., different color or 'Overridden' label)",
        "P1", "Yes"))
    cases.append(c("5.2-008", "5.2", "Grading Workbench", "Add teacher comment to submission",
        "TEACHER", "Grading workbench open",
        "1. Scroll to the Comments section\n2. Type a comment in the text field\n3. Click 'Add Comment' or press Enter",
        "1. The comment appears in the comments list with your name and timestamp\n2. The comment is visible to the student in their feedback view",
        "P1", "Yes"))
    cases.append(c("5.2-009", "5.2", "Grading Workbench", "Edit a teacher comment",
        "TEACHER", "A teacher comment exists on the submission",
        "1. Click the Edit icon on your comment\n2. Modify the text\n3. Save",
        "The comment text updates. Timestamp may show 'Edited'.",
        "P1", ""))
    cases.append(c("5.2-010", "5.2", "Grading Workbench", "Delete a teacher comment",
        "TEACHER", "A teacher comment exists",
        "1. Click the Delete icon on your comment\n2. Confirm",
        "The comment is removed from the comments list.",
        "P1", ""))
    cases.append(c("5.2-011", "5.2", "Grading Workbench", "Finalize grading — mark as complete",
        "TEACHER", "All feedback items reviewed, score set",
        "1. Click 'Finalize' or 'Mark as Complete'\n2. Confirm",
        "1. A completion animation plays (stamp animation)\n2. The submission is marked as graded\n3. The student can now view their feedback\n4. Grading queue updates (submission removed from pending)",
        "P1", "Yes"))
    cases.append(c("5.2-012", "5.2", "Grading Workbench", "Breather card appears after 15-20 submissions",
        "TEACHER", "Has graded 15+ submissions in a session",
        "1. Continue grading submissions one after another\n2. After approximately 15-20 submissions, observe",
        "A 'breather card' appears with a motivational message and grading session stats (graded count, approved count). Encourages taking a break.",
        "P2", ""))
    cases.append(c("5.2-013", "5.2", "Grading Workbench", "Mobile: stacked vertical layout",
        "TEACHER", "On a mobile device (<768px)",
        "1. Open a submission in the Grading workbench on mobile",
        "Student work and AI feedback are stacked vertically instead of side-by-side. All controls remain accessible.",
        "P1", ""))
    cases.append(c("5.2-014", "5.2", "Grading Workbench", "Band score summary card",
        "TEACHER", "Viewing a Writing submission with AI analysis",
        "1. Look at the top of the feedback panel",
        "A summary card shows: Overall band score + individual criterion scores (Task Achievement, Coherence & Cohesion, Lexical Resource, Grammatical Range & Accuracy).",
        "P1", ""))

    # Grading Queue
    cases.append(c("5.3-001", "5.3", "Grading Queue", "View grading queue with pending submissions",
        "TEACHER", "Students have submitted exercises",
        "1. Navigate to the Grading page",
        "A list of pending submissions appears. Each item shows: Student name, Exercise title, Skill badge, Submission date, AI analysis status.",
        "P1", "Yes"))
    cases.append(c("5.3-002", "5.3", "Grading Queue", "Filter grading queue",
        "TEACHER", "Multiple submissions from different classes/skills",
        "1. Use filter options on the Grading page",
        "Queue can be filtered by class, skill, or status. Results update accordingly.",
        "P1", ""))
    cases.append(c("5.3-003", "5.3", "Grading Queue", "Click submission to open workbench",
        "TEACHER", "Pending submissions in the queue",
        "1. Click on a submission in the queue list",
        "The grading workbench opens with the selected submission loaded (split-screen view for Writing/Speaking, or auto-graded results for Reading/Listening).",
        "P1", ""))
    cases.append(c("5.3-004", "5.3", "Grading Queue", "Graded submission removed from pending queue",
        "TEACHER", "Just finalized grading on a submission",
        "1. After finalizing, navigate back to the Grading queue list",
        "The submission you just graded no longer appears in the pending queue.",
        "P1", ""))
    cases.append(c("5.3-005", "5.3", "Grading Queue", "Owner can access grading for all classes",
        "OWNER", "Submissions exist across multiple classes with different teachers",
        "1. Navigate to Grading as Owner",
        "All submissions across all classes and teachers are visible in the queue.",
        "P1", "Yes"))
    cases.append(c("5.3-006", "5.3", "Grading Queue", "Teacher only sees submissions for their classes",
        "TEACHER", "Submissions exist for classes assigned to this teacher and other teachers",
        "1. Navigate to Grading as Teacher",
        "Only submissions for this teacher's assigned classes appear. Other teachers' class submissions are not visible.",
        "P1", "Yes"))

    # Student Feedback View
    cases.append(c("5.4-001", "5.4", "Student Feedback", "Student views graded feedback",
        "STUDENT", "A submission has been finalized/graded by teacher",
        "1. Navigate to the Dashboard\n2. Find the graded assignment\n3. Click to view feedback",
        "1. Feedback page opens showing: overall score, per-criterion band scores\n2. AI feedback items (grammar highlights, vocabulary notes) are visible\n3. Teacher comments are shown with timestamps",
        "P1", "Yes"))
    cases.append(c("5.4-002", "5.4", "Student Feedback", "Student sees score in sticky header",
        "STUDENT", "Viewing feedback page",
        "1. Scroll down the feedback page",
        "The overall score remains visible in a sticky header as you scroll through the detailed feedback.",
        "P1", ""))
    cases.append(c("5.4-003", "5.4", "Student Feedback", "View submission history (multiple attempts)",
        "STUDENT", "Has submitted the same assignment multiple times (if retries allowed)",
        "1. On the feedback page, look for submission history panel",
        "A history panel shows previous attempts with dates and scores. Clicking a previous attempt shows that submission's feedback.",
        "P2", ""))
    cases.append(c("5.4-004", "5.4", "Student Feedback", "Student cannot view ungraded submission feedback",
        "STUDENT", "A submission is pending grading",
        "1. Look for the pending submission on the Dashboard",
        "The assignment shows status 'Submitted' or 'Pending Review' with no feedback link. Feedback is not accessible until grading is finalized.",
        "P1", ""))

    return cases


# ──────────────────────────────────────────────
# EPIC 6 — Student Health
# ──────────────────────────────────────────────
def epic6():
    cases = []
    cases.append(c("6.1-001", "6.1", "Traffic Light Dashboard", "Student cards show color-coded health status",
        "OWNER", "Students exist with varied attendance and assignment completion",
        "1. Navigate to the Dashboard (or Students page)\n2. Observe the student health cards",
        "1. Student cards display with color-coded status: Red (At-Risk), Yellow (Warning), Green (On-Track)\n2. Cards are sorted with at-risk students first, then warning, then on-track\n3. Each card shows: Student name, avatar, health status indicator, attendance %, assignment completion %",
        "P1", "Yes"))
    cases.append(c("6.1-002", "6.1", "Traffic Light Dashboard", "At-Risk (Red) — attendance below 80%",
        "OWNER", "A student has attendance below 80%",
        "1. Navigate to Dashboard/Students\n2. Find the student with low attendance",
        "The student's card is red (At-Risk). Attendance percentage is displayed and is below 80%.",
        "P1", ""))
    cases.append(c("6.1-003", "6.1", "Traffic Light Dashboard", "Warning (Yellow) — attendance 80-90%",
        "OWNER", "A student has attendance between 80-90%",
        "1. Find the student with medium attendance",
        "The student's card is yellow (Warning). Attendance percentage is shown between 80-90%.",
        "P1", ""))
    cases.append(c("6.1-004", "6.1", "Traffic Light Dashboard", "On-Track (Green) — attendance above 90%",
        "OWNER", "A student has attendance above 90%",
        "1. Find the student with high attendance",
        "The student's card is green (On-Track). Attendance percentage is above 90%.",
        "P1", ""))
    cases.append(c("6.1-005", "6.1", "Traffic Light Dashboard", "At-Risk (Red) — assignment completion below 50%",
        "OWNER", "A student has completed less than 50% of assignments",
        "1. Find the student with low completion",
        "The student's card is red (At-Risk). Assignment completion percentage is below 50%.",
        "P1", ""))
    cases.append(c("6.1-006", "6.1", "Traffic Light Dashboard", "Warning (Yellow) — assignment completion 50-75%",
        "OWNER", "A student has 50-75% completion",
        "1. Find the student",
        "Yellow card. Completion percentage shown between 50-75%.",
        "P1", ""))
    cases.append(c("6.1-007", "6.1", "Traffic Light Dashboard", "On-Track (Green) — assignment completion above 75%",
        "OWNER", "A student has >75% completion",
        "1. Find the student",
        "Green card. Completion percentage above 75%.",
        "P1", ""))
    cases.append(c("6.1-008", "6.1", "Traffic Light Dashboard", "Overall status uses worst metric",
        "OWNER", "A student has >90% attendance but <50% assignment completion",
        "1. Find this student on the dashboard",
        "Despite good attendance, the card is Red (At-Risk) because assignment completion is the worse metric.",
        "P1", "Yes"))
    cases.append(c("6.1-009", "6.1", "Traffic Light Dashboard", "New student with no data defaults to Green",
        "OWNER", "A newly added student with no attendance or assignments",
        "1. Find the new student on the dashboard",
        "The student's card is Green (On-Track) by default.",
        "P2", ""))
    cases.append(c("6.1-010", "6.1", "Traffic Light Dashboard", "Filter students by health status",
        "OWNER", "Students with different health statuses exist",
        "1. On the Dashboard, click one of the status filters in the health summary bar (e.g., 'At-Risk')",
        "Only students with the selected status are shown. Other students are hidden.",
        "P1", ""))
    cases.append(c("6.1-011", "6.1", "Traffic Light Dashboard", "Filter students by class",
        "OWNER", "Students in multiple classes",
        "1. Use the Class filter dropdown",
        "Only students in the selected class are shown.",
        "P1", ""))
    cases.append(c("6.1-012", "6.1", "Traffic Light Dashboard", "Search students by name",
        "OWNER", "Multiple students exist",
        "1. Type a student's name in the search field",
        "Student cards filter to show only students matching the search text.",
        "P1", ""))
    cases.append(c("6.1-013", "6.1", "Traffic Light Dashboard", "Click student card opens profile overlay",
        "OWNER", "Student cards are displayed",
        "1. Click on a student's card",
        "A profile overlay or detail panel opens showing the student's detailed information: full health metrics, class enrollment, recent attendance, assignment history.",
        "P1", "Yes"))
    cases.append(c("6.1-014", "6.1", "Traffic Light Dashboard", "Health summary bar shows counts",
        "OWNER", "Students exist with various statuses",
        "1. Look at the health summary bar at the top of the dashboard",
        "The summary bar shows counts for each status: e.g., 'Engaged: 15 | At-Risk: 3 | Inactive: 2'. Counts match the visible cards.",
        "P1", ""))
    cases.append(c("6.1-015", "6.1", "Traffic Light Dashboard", "Teacher sees health dashboard for their students",
        "TEACHER", "Teacher is assigned to classes with students",
        "1. Log in as Teacher\n2. Navigate to Dashboard or Students",
        "1. The health dashboard shows only students in the teacher's assigned classes\n2. Student cards with health indicators are displayed",
        "P1", "Yes"))
    cases.append(c("6.1-016", "6.1", "Traffic Light Dashboard", "Dashboard loads within 1 second",
        "OWNER", "Center has 50+ students",
        "1. Navigate to the Dashboard\n2. Measure load time",
        "Student health data loads and renders within approximately 1 second. Skeleton loading state may briefly appear before data populates.",
        "P2", ""))
    cases.append(c("6.1-017", "6.1", "Traffic Light Dashboard", "Student cannot see health dashboard",
        "STUDENT", "Logged in as Student",
        "1. Navigate to Dashboard",
        "Student sees their own assignment dashboard (grouped by urgency), NOT the student health view. No other student's data is visible.",
        "P1", "Yes"))

    # Interventions / flags
    cases.append(c("6.2-001", "6.2", "Interventions", "Flag a student for intervention",
        "TEACHER", "Student health view visible",
        "1. Click on an at-risk student's card\n2. In the profile overlay, click 'Flag for Intervention' or similar action\n3. Add a note about the concern",
        "1. The student gets a flag icon on their card\n2. The flag note is saved and visible to other teachers/admin",
        "P2", ""))
    cases.append(c("6.2-002", "6.2", "Interventions", "View flagged students",
        "OWNER", "Students have been flagged",
        "1. Look for a filter or section showing flagged students",
        "Flagged students are identifiable by a flag icon or can be filtered to show only flagged students.",
        "P2", ""))

    return cases


# ──────────────────────────────────────────────
# EPIC 7 — Notifications
# ──────────────────────────────────────────────
def epic7():
    cases = []
    cases.append(c("7.1-001", "7.1", "Engagement Emails", "7-day streak triggers celebration email",
        "STUDENT", "Student has been graded on exercises for 7 consecutive days",
        "1. Ensure student submits and gets graded for 7 consecutive days\n2. Check the student's email inbox after the 7th day",
        "Student receives a celebration/streak email (e.g., '7-day streak! Keep it up!'). If parent emails are configured, parents also receive the email.",
        "P1", ""))
    cases.append(c("7.1-002", "7.1", "Engagement Emails", "No streak email with a gap day",
        "STUDENT", "Student has 6 graded days with a gap (e.g., Mon-Wed, then Fri-Sat = 6 days but not consecutive)",
        "1. Check email inbox",
        "No streak email is received. The streak requires 7 consecutive calendar days.",
        "P2", ""))
    cases.append(c("7.1-003", "7.1", "Engagement Emails", "Personal best triggers congratulation email",
        "STUDENT", "Student has previous Writing/Speaking scores and just received a new highest score",
        "1. After receiving a score higher than all previous scores for the same skill\n2. Check email inbox",
        "Student receives a 'Personal Best' congratulation email highlighting the new high score.",
        "P1", ""))
    cases.append(c("7.1-004", "7.1", "Engagement Emails", "No personal best email on first submission",
        "STUDENT", "Student submits their very first exercise (no comparison scores)",
        "1. Submit first exercise\n2. Wait for grading\n3. Check email",
        "No 'Personal Best' email is sent for the first submission (there's nothing to compare against).",
        "P2", ""))
    cases.append(c("7.1-005", "7.1", "Engagement Emails", "Max 1 engagement email per student per day",
        "STUDENT", "Student triggers both a streak and a personal best on the same day",
        "1. Check email inbox",
        "Student receives a single combined email (not two separate emails) covering both the streak and personal best.",
        "P1", ""))
    cases.append(c("7.1-006", "7.1", "Engagement Emails", "Parent email notification",
        "STUDENT", "Student has parent emails configured in profile",
        "1. Trigger an engagement event (streak or personal best)\n2. Check parent email inbox",
        "Parents receive the same engagement email that was sent to the student.",
        "P1", ""))

    # Notification preferences
    cases.append(c("7.2-001", "7.2", "Notification Preferences", "Student can manage email preferences",
        "STUDENT", "Logged in as Student",
        "1. Navigate to My Profile\n2. Find the notification/email preferences section\n3. Toggle off streak emails\n4. Save",
        "1. Preference is saved\n2. After the change, streak emails are no longer sent to this student",
        "P1", ""))
    cases.append(c("7.2-002", "7.2", "Notification Preferences", "Add parent email in profile",
        "STUDENT", "Logged in as Student",
        "1. Navigate to My Profile\n2. Find the parent email section\n3. Add a parent's email address\n4. Save",
        "1. Parent email is saved and visible on the profile\n2. Future engagement emails are also sent to this parent email",
        "P1", ""))
    cases.append(c("7.2-003", "7.2", "Notification Preferences", "Remove parent email",
        "STUDENT", "A parent email is configured",
        "1. Navigate to My Profile\n2. Remove the parent email\n3. Save",
        "Parent email is removed. Future emails are no longer sent to that address.",
        "P2", ""))

    # In-app notifications
    cases.append(c("7.3-001", "7.3", "In-App Notifications", "Notification bell shows unread count",
        "ALL", "Logged in; notifications exist (e.g., schedule change)",
        "1. Look at the notification bell icon in the top bar",
        "The bell icon shows a badge with the unread notification count (e.g., red circle with '2').",
        "P1", "Yes"))
    cases.append(c("7.3-002", "7.3", "In-App Notifications", "Click bell to view notifications",
        "ALL", "Unread notifications exist",
        "1. Click the notification bell icon",
        "A dropdown or panel shows the list of notifications with titles, descriptions, and timestamps. Unread notifications are visually distinct.",
        "P1", ""))
    cases.append(c("7.3-003", "7.3", "In-App Notifications", "Mark notification as read",
        "ALL", "An unread notification exists",
        "1. Click on an unread notification or click 'Mark as read'",
        "The notification changes to read state (visual change). The unread count decreases.",
        "P2", ""))
    cases.append(c("7.3-004", "7.3", "In-App Notifications", "Schedule change notification received",
        "STUDENT", "An Owner has moved a session that includes this student's class",
        "1. Check the notification bell",
        "A notification about the schedule change appears (e.g., 'Class session moved to [new time]').",
        "P1", ""))
    cases.append(c("7.3-005", "7.3", "In-App Notifications", "Assignment notification received",
        "STUDENT", "A teacher has created a new assignment for the student's class",
        "1. Check the notification bell",
        "A notification about the new assignment appears (e.g., 'New assignment: [title]').",
        "P1", ""))
    cases.append(c("7.3-006", "7.3", "In-App Notifications", "Grading complete notification for student",
        "STUDENT", "Teacher has finalized grading on the student's submission",
        "1. Check the notification bell",
        "A notification appears (e.g., 'Your [exercise title] has been graded. View feedback.').",
        "P1", ""))

    return cases


# ──────────────────────────────────────────────
# EPIC 9 — Billing
# ──────────────────────────────────────────────
def epic9():
    cases = []
    # Story 9.1 — Billing Dashboard
    cases.append(c("9.1-001", "9.1", "Billing Dashboard", "Billing page shows tier, students, cost, next billing date",
        "OWNER", "Logged in as Owner; center has an active subscription",
        "1. Navigate to Settings > Billing",
        "The billing page displays four metric cards: Current Tier name, Enrolled Student count, Monthly Cost estimate, and Next Billing Date.",
        "P1", "Yes"))
    cases.append(c("9.1-002", "9.1", "Billing Dashboard", "Pilot phase shows free status",
        "OWNER", "Center is in the free pilot phase",
        "1. Navigate to Settings > Billing",
        "1. Tier shows 'Free Pilot'\n2. Monthly cost shows '$0.00'\n3. Next billing date shows 'N/A'\n4. Subscription status badge shows 'Pilot'",
        "P1", ""))
    cases.append(c("9.1-003", "9.1", "Billing Dashboard", "Payment history table (empty during pilot)",
        "OWNER", "Center is in pilot phase",
        "1. On the Billing page, scroll to the payment history section",
        "Payment history table shows 'No payments yet' or an empty state message.",
        "P1", ""))
    cases.append(c("9.1-004", "9.1", "Billing Dashboard", "Usage chart shows enrolled students over 6 months",
        "OWNER", "Center has been active for multiple months",
        "1. On the Billing page, look at the usage chart",
        "A bar chart displays enrolled student counts for each of the last 6 months. Bars are labeled with month names and counts.",
        "P2", ""))
    cases.append(c("9.1-005", "9.1", "Billing Dashboard", "'Manage Subscription' disabled during pilot",
        "OWNER", "Center is in pilot phase",
        "1. Look for the 'Manage Subscription' or 'Change Plan' button",
        "The button is disabled/grayed out with text like 'Free during pilot'. Clicking it does nothing.",
        "P1", ""))
    cases.append(c("9.1-006", "9.1", "Billing Dashboard", "Active subscription status badge",
        "OWNER", "Center has an active paid subscription",
        "1. Navigate to Settings > Billing",
        "Subscription status badge shows 'Active' in green. All metric cards show current values.",
        "P1", ""))
    cases.append(c("9.1-007", "9.1", "Billing Dashboard", "Past Due status badge",
        "OWNER", "Subscription payment has failed",
        "1. Navigate to Settings > Billing",
        "Status badge shows 'Past Due' in red/orange. An alert or message may indicate payment needs to be updated.",
        "P1", ""))
    cases.append(c("9.1-008", "9.1", "Billing Dashboard", "Grace Period status and banner",
        "OWNER", "Subscription is in grace period",
        "1. Navigate to Settings > Billing",
        "1. Status badge shows 'Grace Period' in yellow/amber\n2. A banner appears indicating the grace period deadline\n3. 'Update Payment Method' action is prominently displayed",
        "P1", ""))
    cases.append(c("9.1-009", "9.1", "Billing Dashboard", "Payment history with receipts",
        "OWNER", "Center has completed payments",
        "1. Check the payment history table",
        "Table shows: Date, Amount, Status (e.g., Paid), and a Receipt link for each payment. Clicking Receipt opens or downloads the receipt.",
        "P1", ""))
    cases.append(c("9.1-010", "9.1", "Billing Dashboard", "Only Owner can access Billing page",
        "ADMIN", "Logged in as Admin",
        "1. Navigate to Settings > Billing",
        "The Billing tab may not be visible, or accessing the URL redirects to dashboard/shows access denied.",
        "P1", "Yes"))

    # Story 9.4 — Tier Management
    cases.append(c("9.4-001", "9.4", "Tier Management", "Tier comparison table shows all tiers",
        "OWNER", "On the Billing page",
        "1. Navigate to Settings > Billing\n2. Scroll to the tier comparison section",
        "A table displays all tiers: Starter ($20/mo, 30 students), Growth ($50/mo, 100 students), Enterprise ($100/mo, unlimited). Each tier shows its features and pricing.",
        "P1", ""))
    cases.append(c("9.4-002", "9.4", "Tier Management", "Current tier is highlighted in comparison table",
        "OWNER", "Center has an active subscription",
        "1. Look at the tier comparison table",
        "The current tier is visually highlighted (e.g., bordered, different background color, 'Current Plan' label).",
        "P1", ""))
    cases.append(c("9.4-003", "9.4", "Tier Management", "Student limit warning badge on smaller tiers",
        "OWNER", "Center has 50 students (exceeds Starter's 30 limit)",
        "1. Look at the Starter tier in the comparison table",
        "A warning badge indicates the center exceeds this tier's student limit (e.g., 'Exceeds limit: 50/30 students').",
        "P1", ""))
    cases.append(c("9.4-004", "9.4", "Tier Management", "Upgrade shows 'Immediate (prorated)' timing",
        "OWNER", "Center is on Starter tier",
        "1. Look at the Growth tier\n2. Hover over or note the upgrade button",
        "The upgrade option shows timing text: 'Immediate (prorated)' indicating the change takes effect right away with prorated billing.",
        "P1", ""))
    cases.append(c("9.4-005", "9.4", "Tier Management", "Downgrade shows 'Effective next billing cycle' timing",
        "OWNER", "Center is on Growth tier",
        "1. Look at the Starter tier\n2. Note the downgrade option",
        "The downgrade option shows timing text: 'Effective next billing cycle'.",
        "P1", ""))
    cases.append(c("9.4-006", "9.4", "Tier Management", "Downgrade confirmation warns about student count",
        "OWNER", "Center is on Growth (100 students) and has 50 students enrolled",
        "1. Click 'Downgrade' to Starter (30 students limit)\n2. Observe the confirmation dialog",
        "A confirmation dialog warns: 'Your center currently has 50 students. The Starter tier allows a maximum of 30 students. You will need to reduce your student count.' The warning is clearly visible before confirming.",
        "P1", "Yes"))
    cases.append(c("9.4-007", "9.4", "Tier Management", "Pilot users see 'Subscribe' buttons per tier",
        "OWNER", "Center is in pilot phase",
        "1. Look at the tier comparison table",
        "Each tier shows a 'Subscribe' button (instead of 'Change Plan'). Clicking subscribes the center to that tier.",
        "P1", ""))
    cases.append(c("9.4-008", "9.4", "Tier Management", "Change Plan button opens Polar portal",
        "OWNER", "Center has an active subscription",
        "1. Click 'Change Plan' or 'Manage Subscription'",
        "A new browser window/tab opens with the Polar.sh customer portal where payment and plan management can be completed.",
        "P1", ""))
    cases.append(c("9.4-009", "9.4", "Tier Management", "After upgrading, billing page reflects new tier",
        "OWNER", "Just completed an upgrade via Polar",
        "1. Return to the Billing page\n2. Refresh the page",
        "1. Tier name updates to the new plan\n2. Monthly cost updates\n3. Student limit reflects the new tier\n4. Success message or updated status badge appears",
        "P1", ""))
    cases.append(c("9.4-010", "9.4", "Tier Management", "Non-Owner role cannot see or access billing",
        "TEACHER", "Logged in as Teacher",
        "1. Check the sidebar for Settings\n2. Try to access the Billing URL directly",
        "Settings is not visible in sidebar. Direct URL access to billing redirects to dashboard or shows access denied.",
        "P1", "Yes"))

    return cases


# ──────────────────────────────────────────────
# EPIC 10 — Marketing Website
# ──────────────────────────────────────────────
def epic10():
    cases = []
    cases.append(c("10.1-001", "10.1", "Landing Page", "Single scrollable page loads",
        "PUBLIC", "",
        "1. Open classlite.app in a browser",
        "A single scrollable page loads with all sections visible as you scroll. There are no additional page routes.",
        "P1", "Yes"))
    cases.append(c("10.1-002", "10.1", "Landing Page", "Nav: logo and 'Sign In' link",
        "PUBLIC", "",
        "1. Look at the top navigation bar",
        "1. ClassLite logo/wordmark is on the left\n2. 'Sign In' link is on the right\n3. Clicking 'Sign In' navigates to the app sign-up page (my.classlite.app/sign-up or similar)",
        "P1", ""))
    cases.append(c("10.1-003", "10.1", "Landing Page", "Hero section content",
        "PUBLIC", "",
        "1. Look at the hero section (first full-screen section)",
        "1. Headline: 'Teaching, without the clutter.' (or similar)\n2. Subheadline describing the LMS for IELTS centers\n3. 'Get Started' CTA button\n4. CTA button links to the sign-up page",
        "P1", ""))
    cases.append(c("10.1-004", "10.1", "Landing Page", "Three value cards",
        "PUBLIC", "",
        "1. Scroll to the value proposition section",
        "Three cards are displayed with headlines: 'Grade in minutes', 'Know who needs help', 'Schedule without the mess' (or similar). Each has descriptive text.",
        "P1", ""))
    cases.append(c("10.1-005", "10.1", "Landing Page", "How It Works section",
        "PUBLIC", "",
        "1. Scroll to the 'How It Works' section",
        "Three steps are displayed with icons/illustrations: 1. Build exercises, 2. Students submit, 3. AI helps grade (or similar wording).",
        "P1", ""))
    cases.append(c("10.1-006", "10.1", "Landing Page", "Footer content",
        "PUBLIC", "",
        "1. Scroll to the bottom of the page",
        "Footer displays: '© 2026 ClassLite' (or current year).",
        "P1", ""))

    # Responsive & Performance
    cases.append(c("10.1-007", "10.1", "Responsive", "Mobile responsive (375px)",
        "PUBLIC", "",
        "1. Open classlite.app on a 375px wide screen (iPhone SE size)\n2. Scroll through all sections",
        "1. All content is readable without horizontal scrolling\n2. Images and cards stack vertically\n3. Navigation remains accessible\n4. No content overflow or clipping",
        "P1", "Yes"))
    cases.append(c("10.1-008", "10.1", "Responsive", "Tablet responsive (768px)",
        "PUBLIC", "",
        "1. Open classlite.app on a 768px wide screen\n2. Scroll through",
        "Layout adapts for tablet: cards may show in 2-column layout. All content is accessible and properly sized.",
        "P1", ""))
    cases.append(c("10.1-009", "10.1", "Responsive", "Desktop layout (1024px+)",
        "PUBLIC", "",
        "1. Open classlite.app on desktop (1024px+)",
        "Full desktop layout with proper spacing, multi-column cards, and maximum content width.",
        "P1", ""))
    cases.append(c("10.1-010", "10.1", "SEO", "Page title and meta description",
        "PUBLIC", "",
        "1. Open classlite.app\n2. Check the browser tab title\n3. View page source or use SEO inspector",
        "1. Page has a descriptive title (e.g., 'ClassLite — Lightweight LMS for IELTS Centers')\n2. Meta description tag exists with relevant content\n3. Open Graph tags exist for social sharing",
        "P1", ""))
    cases.append(c("10.1-011", "10.1", "SEO", "Favicon is present",
        "PUBLIC", "",
        "1. Open classlite.app\n2. Look at the browser tab icon",
        "A custom favicon (ClassLite icon) appears in the browser tab, not the default browser icon.",
        "P2", ""))
    cases.append(c("10.1-012", "10.1", "Performance", "Page loads within 3 seconds",
        "PUBLIC", "",
        "1. Open classlite.app on a standard internet connection\n2. Observe load time",
        "The page is fully rendered and interactive within approximately 3 seconds. No long loading spinners or blank screens.",
        "P1", ""))
    cases.append(c("10.1-013", "10.1", "Brand", "Brand palette: Royal Blue, Golden Yellow, white, dark navy",
        "PUBLIC", "",
        "1. Observe the overall color scheme",
        "Primary color is Royal Blue (#2563EB). Accent colors include Golden Yellow. Background is white. Text is dark navy. Colors are consistent throughout.",
        "P2", ""))
    cases.append(c("10.1-014", "10.1", "Brand", "Typography: Outfit headings, Inter body",
        "PUBLIC", "",
        "1. Inspect the fonts used (browser dev tools or visual inspection)",
        "Headings use the Outfit font family. Body text uses the Inter font family. Fonts are consistent across the page.",
        "P3", ""))

    return cases


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets = [
        ("Epic 1 - Tenant & Users", epic1()),
        ("Epic 2 - Scheduling", epic2()),
        ("Epic 3 - Exercise Builder", epic3()),
        ("Epic 4 - Submissions", epic4()),
        ("Epic 5 - AI Grading", epic5()),
        ("Epic 6 - Student Health", epic6()),
        ("Epic 7 - Notifications", epic7()),
        ("Epic 9 - Billing", epic9()),
        ("Epic 10 - Marketing", epic10()),
    ]

    total = 0
    for name, cases in sheets:
        add_sheet(wb, name, cases)
        print(f"{name}: {len(cases)} cases")
        total += len(cases)

    print(f"\nTotal: {total} cases")

    out = "/Users/ducdo/Desktop/code/personal/classlite/_bmad-output/test-cases/classlite-manual-test-cases.xlsx"
    wb.save(out)
    print(f"\nSaved to {out}")


if __name__ == "__main__":
    main()
